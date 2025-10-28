import dash
from dash import dcc, html, Input, Output, State, dash_table, no_update
import plotly.express as px
import plotly.graph_objs as go
import pandas as pd
import os
from dash import ctx
import io
import dash_bootstrap_components as dbc
import logging
import glob
import numpy as np
import xlsxwriter
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile
import requests
from io import BytesIO
import pyarrow
import pyarrow.parquet as pq
import polars as pl
import traceback
import base64
from dash.exceptions import PreventUpdate
from data_uploader import process_new_file, read_excel_file, safe_filename
from github import Github, InputGitTreeElement
import uuid
import time
import threading
from datetime import datetime, timedelta, date
import subprocess
import psycopg2
from sqlalchemy import create_engine, text
from collections import defaultdict
from sqlalchemy.exc import OperationalError
import re
from difflib import get_close_matches
from typing import Optional
# --------------------
# НАСТРОЙКИ
# --------------------
HEIGHT_PER_BAR = 30  # высота одной строки в px
MAX_VISIBLE_BARS = 50  # сколько строк показывать без прокрутки
MAX_HEIGHT = HEIGHT_PER_BAR * MAX_VISIBLE_BARS  # высота контейнера в px

# --------------------
# ЗАГРУЗКА И ПРЕДОБРАБОТКА (один раз при старте)
# --------------------
def safe_read_excel(path):
    try:
        if path and os.path.exists(path):
            return pd.read_excel(path)
    except Exception:
        pass
    return pd.DataFrame()

df_result = safe_read_excel('итог_по_месяцу.xlsx')
df_fast = safe_read_excel('самые_ходовые.xlsx')
df_restock = safe_read_excel('чаще_всего_пополнялись.xlsx')
df_peaks = pd.read_excel('всплески_продаж1.xlsx')
df_peaks['Дата'] = pd.to_datetime(df_peaks['Дата'])

# Опционально: привести колонку Всплеск к булевому типу, если нужно
df_peaks['Всплеск'] = df_peaks['Всплеск'].astype(bool)

# Приведение числовых колонок
if not df_fast.empty:
    df_fast['Всего_продано'] = pd.to_numeric(df_fast.get('Всего_продано', 0), errors='coerce').fillna(0)
    df_fast = df_fast.dropna(subset=['Номенклатура'])

if not df_restock.empty:
    df_restock['Всего_пополнено'] = pd.to_numeric(df_restock.get('Всего_пополнено', df_restock.get('Всего_продано', 0)), errors='coerce').fillna(0)
    df_restock = df_restock.dropna(subset=['Номенклатура'])


# Группировки для топов
fast_grouped = df_fast.groupby(['Склад', 'Номенклатура', 'Артикул'], as_index=False)['Всего_продано'].sum() if not df_fast.empty else pd.DataFrame()
restock_grouped = df_restock.groupby(['Склад', 'Номенклатура', 'Артикул'], as_index=False)['Всего_пополнено'].sum() if not df_restock.empty else pd.DataFrame()

# Уникальные значения для фильтров
unique_sklads = df_result['Склад'].dropna().unique().tolist() if not df_result.empty else []
unique_peak_sklads = sorted(df_peaks['Склад'].dropna().unique()) if not df_peaks.empty else []
unique_peak_articles = sorted(df_peaks['Артикул'].dropna().unique()) if not df_peaks.empty else []
unique_peak_noms = sorted(df_peaks['Номенклатура'].dropna().unique()) if not df_peaks.empty else []

# --- Подключение к Railway ---
DB_URL = "postgresql://postgres:SyngvjjliGqUBYDKibMmoOWCVUZVdFjc@tramway.proxy.rlwy.net:13502/railway"
engine = create_engine(DB_URL, pool_pre_ping=True)
TABLE_NAME = "alyans_refresh_v3"

# -------------------- Глобальный буфер --------------------
upload_df = None
# -------------------- Вспомогательные функции --------------------
def clean_articles(df: pd.DataFrame) -> pd.DataFrame:
    """Минимальная очистка: убираем пробелы, приводим к строкам"""
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]
    str_cols = df.select_dtypes(include="object").columns
    for c in str_cols:
        df[c] = df[c].str.strip()
    return df

def calc_changes(df_new: pd.DataFrame, df_prev: pd.DataFrame) -> pd.DataFrame:
    """Расчет продано и пополнено"""
    df = df_new.merge(df_prev[["товар_id", "склад", "остаток"]], 
                      on=["товар_id", "склад"], how="left", suffixes=("", "_вчера"))
    df["продано"] = (df["остаток_вчера"] - df["остаток"]).clip(lower=0).fillna(0).astype(int)
    df["пополнение"] = (df["остаток"] - df["остаток_вчера"]).clip(lower=0).fillna(0).astype(int)
    df.drop(columns=["остаток_вчера"], inplace=True)
    return df

def get_last_date_in_db() -> Optional[date]:
    """Возвращает последнюю дату из таблицы"""
    with engine.connect() as conn:
        last_date = conn.execute(text(f"SELECT MAX(дата) FROM {TABLE_NAME}")).scalar()
    if last_date is None:
        return None
    return pd.to_datetime(last_date).date()

def extract_date_from_filename(filename: str) -> datetime.date | None:
    """Парсит дату из имени файла вида 'dd.mm.xlsx'"""
    match = re.search(r'(\d{1,2})\.(\d{1,2})', filename)
    if not match:
        return None
    day, month = map(int, match.groups())
    return datetime(2025, month, day).date()  # год можно менять

def save_to_db(df: pd.DataFrame, chunk_size: int = 50000):
    """Генератор по частичной записи в базу"""
    total_chunks = math.ceil(len(df) / chunk_size)
    with engine.begin() as conn:
        for i in range(total_chunks):
            start = i * chunk_size
            end = min(start + chunk_size, len(df))
            df.iloc[start:end].to_sql(TABLE_NAME, conn, if_exists="append", index=False)
            yield int((i + 1) / total_chunks * 100)



def _ensure_list(value):
    """Гарантирует, что значение является списком."""
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def get_latest_upload_date():
    """Возвращает дату последнего загруженного файла в формате ДД.ММ.ГГГГ"""
    upload_dir = os.path.join("data", "new_uploads")
    pattern = os.path.join(upload_dir, "new_uploads_*.csv")
    files = glob.glob(pattern)
    if not files:
        return "нет данных"

    latest_date = None
    for f in files:
        filename = os.path.basename(f)
        # пример: new_uploads_2025-10-02_14-46.csv
        try:
            date_part = filename.split("_")[2]  # '2025-10-02'
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            if latest_date is None or dt > latest_date:
                latest_date = dt
        except Exception:
            continue

    if latest_date:
        return latest_date.strftime("%d.%m.%Y")
    else:
        return "неизвестно"


# ---------- Фильтры ----------
def get_unique_sklads():
    try:
        sql = text("SELECT DISTINCT склад FROM alyans_refresh_v3 WHERE склад IS NOT NULL ORDER BY склад;")
        with engine.connect() as conn:
            rows = conn.execute(sql).fetchall()
        return [r[0] for r in rows]
    except Exception:
        logger.exception("[get_unique_sklads] Ошибка")
        return []


def get_unique_groups():
    try:
        sql = text("SELECT DISTINCT группа FROM alyans_refresh_v3 WHERE группа IS NOT NULL ORDER BY группа;")
        with engine.connect() as conn:
            rows = conn.execute(sql).fetchall()
        return [r[0] for r in rows]
    except Exception:
        logger.exception("[get_unique_groups] Ошибка")
        return []


# ---------- ТОП товаров ----------
def get_top_products(top_n=100, sklads=None, groups=None, project=None):
    """
    Возвращает ТОП товаров по продажам с фильтрами:
    - Склад
    - Группа
    - Проект (Корея / Китай)
    Оптимизировано для больших данных и предотвращения OOM на Railway.
    """
    sklads = _ensure_list(sklads)
    groups = _ensure_list(groups)
    params = {"top_n": int(top_n)}
    where = ["1=1"]

    # --- 🔹 Фильтр по складам ---
    if sklads:
        where.append("склад = ANY(:sklads)")
        params["sklads"] = sklads

    # --- 🔹 Фильтр по группам ---
    if groups:
        where.append("группа = ANY(:groups)")
        params["groups"] = groups

    # --- 🔹 Проектные группы ---
    korea_groups = [
        "ПРОЕКТ ЭЛЕКТРИКА\\СТАРТВОЛЬТ-ИНОМАРКИ",
        "ПРОЕКТ KOREA ЛЕГКОВЫЕ ОПТ\\MANDO-ЛЕГКОВОЙ ОБЩАЯ\\MANDO-КОНТРОЛЬ",
        "ПРОЕКТ CHINA\\CHINA-РТИ ОБЩАЯ\\CHINA-ПРОКЛАДКИ СИЛ",
        "ПРОЕКТ ИНОМАРКИ ГРУЗОВЫЕ ОПТ\\SAMPA",
        "ПРОЕКТ ИНОМАРКИ ГРУЗОВЫЕ ОПТ\\LUZAR-ИНОМАРКИ ГРУЗОВЫЕ",
        "ПРОЕКТ АВТОКОМПОНЕНТЫ\\PSP",
        "ПРОЕКТ ИНОМАРКИ ЛЕГКОВЫЕ ОПТ\\BOSCH ОБЩАЯ\\BOSCH ИНОМАРКИ ГРУЗ",
        "ПРОЕКТ РОЗНИЦА\\*ГРУППА ИНОМАРКИ ЛЕГКОВЫЕ ОБЩАЯ\\ECO-ИНОМАРКИ",
        "ПРОЕКТ KOREA ГРУЗОВЫЕ ОПТ\\HYUNDAI/KIA-ГРУЗОВОЙ ОБЩАЯ\\MOBIS KOREA-ГРУЗОВОЙ",
        "ПРОЕКТ MEGAPOWER ЗАПЧАСТИ\\MR-РК ТОРМ.НАКЛАДКИ",
        "ПРОЕКТ ЭЛЕКТРИКА\\ПРОЕКТ ЭЛЕКТРОСИЛА ОБЩАЯ\\TESLA-ГЕНЕРАТОРЫ СТАРТЕРА",
    ]

    china_groups = [
        "ПРОЕКТ КАМАЗ ГОРОД\\КИТАЙ-КАМАЗ",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\SHACMAN OE",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\HOWO SITRAK",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\ПОДПРОЕКТ JAC ОБЩАЯ\\JAC-ГРУЗОВОЙ OE",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\FAW OE",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\ПОДПРОЕКТ JAC ОБЩАЯ\\JAC-ЛЕГКОВОЙ OE",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\DONGFENG ОБЩАЯ\\DONGFENG OE",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\FOTON OE",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\MOVELEX-КИТАЙ ОБЩАЯ\\MOVELEX-JAC",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\ПОДПРОЕКТ JAC ОБЩАЯ\\JAC-ЦС",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\MOVELEX-КИТАЙ ОБЩАЯ\\MOVELEX-SHACMAN",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\MOVELEX-КИТАЙ ОБЩАЯ\\MOVELEX-SITRAK",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\ПОДПРОЕКТ JAC ОБЩАЯ\\КАМАЗ КОМПАС",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\+КИТАЙ ГРУЗОВЫЕ ОПТ-УЦЕНКА",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\+КИТАЙ ГРУЗОВЫЕ ОПТ-ЗАКРЫТО",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\CREATEK",
        "ПРОЕКТ МАЗ\\КИТАЙ-МАЗ",
        "ПРОЕКТ ПНЕВМО\\ПНЕВМО-КИТАЙ",
        "ПРОЕКТ КИТАЙ ГРУЗОВЫЕ ОПТ\\WEICHAI",
    ]

    # --- 🔹 Обработка фильтра "Проект" ---
    if project == "Корея":
        params["start_date"] = (datetime.utcnow().date() - timedelta(days=365))
        where.append("дата >= :start_date")
        params["korea_groups"] = korea_groups
        where.append("группа = ANY(:korea_groups)")

    elif project == "Китай":
        params["start_date"] = (datetime.utcnow().date() - timedelta(days=365))
        where.append("дата >= :start_date")
        params["china_groups"] = china_groups
        where.append("группа = ANY(:china_groups)")

    # --- 🔹 Защита от слишком широкого запроса ---
    if not sklads and not groups and not project:
        logger.warning("⚠️ Слишком широкий запрос без фильтров — пропущен.")
        return pd.DataFrame()

    where_clause = " AND ".join(where)

    aggregate = len(sklads) > 1

    try:
        with engine.connect() as conn:
            if aggregate:
                sql = f"""
                    SELECT товар_id, артикул, наименование,
                           SUM(продано) AS всего_продано,
                           SUM(пополнение) AS всего_пополнено
                    FROM alyans_refresh_v3
                    WHERE {where_clause}
                    GROUP BY товар_id, артикул, наименование
                    HAVING SUM(продано) > 0
                    ORDER BY всего_продано DESC
                    LIMIT :top_n
                """
            else:
                sql = f"""
                    SELECT склад, товар_id, артикул, наименование,
                           SUM(продано) AS всего_продано,
                           SUM(пополнение) AS всего_пополнено
                    FROM alyans_refresh_v3
                    WHERE {where_clause}
                    GROUP BY склад, товар_id, артикул, наименование
                    HAVING SUM(продано) > 0
                    ORDER BY всего_продано DESC
                    LIMIT :top_n
                """
            res = conn.execute(text(sql), params)
            df = pd.DataFrame(res.fetchall(), columns=res.keys())

        if df.empty:
            return df

        # --- Приведение типов ---
        for col in ["всего_продано", "всего_пополнено"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

        return df

    except Exception:
        logger.exception("[get_top_products] Ошибка получения ТОП товаров")
        return pd.DataFrame()




# ---------- Временной ряд ----------
def get_product_timeseries(tovar_id=None, sklads=None, project=None):
    """
    Возвращает временной ряд по товару для графика.
    Поддерживает фильтры по складу, проекту и артикулу/товар_id.
    """
    if not tovar_id:
        return pd.DataFrame()

    sklads = _ensure_list(sklads)
    params = {"tovar_id": str(tovar_id)}
    where = ["(товар_id = :tovar_id OR артикул = :tovar_id)"]

    # --- 🔹 Фильтр по складам ---
    if sklads:
        where.append("склад = ANY(:sklads)")
        params["sklads"] = sklads

    # --- 🔹 Проектные группы ---
    korea_groups = [
        "ПРОЕКТ ЭЛЕКТРИКА\\СТАРТВОЛЬТ-ИНОМАРКИ",
        "ПРОЕКТ KOREA ЛЕГКОВЫЕ ОПТ\\MANDO-ЛЕГКОВОЙ ОБЩАЯ\\MANDO-КОНТРОЛЬ",
        "ПРОЕКТ CHINA\\CHINA-РТИ ОБЩАЯ\\CHINA-ПРОКЛАДКИ СИЛ",
        "ПРОЕКТ ИНОМАРКИ ГРУЗОВЫЕ ОПТ\\SAMPA",
        "ПРОЕКТ ИНОМАРКИ ГРУЗОВЫЕ ОПТ\\LUZAR-ИНОМАРКИ ГРУЗОВЫЕ",
        "ПРОЕКТ АВТОКОМПОНЕНТЫ\\PSP",
        "ПРОЕКТ ИНОМАРКИ ЛЕГКОВЫЕ ОПТ\\BOSCH ОБЩАЯ\\BOSCH ИНОМАРКИ ГРУЗ",
        "ПРОЕКТ РОЗНИЦА\\*ГРУППА ИНОМАРКИ ЛЕГКОВЫЕ ОБЩАЯ\\ECO-ИНОМАРКИ",
        "ПРОЕКТ KOREA ГРУЗОВЫЕ ОПТ\\HYUNDAI/KIA-ГРУЗОВОЙ ОБЩАЯ\\MOBIS KOREA-ГРУЗОВОЙ",
        "ПРОЕКТ MEGAPOWER ЗАПЧАСТИ\\MR-РК ТОРМ.НАКЛАДКИ",
        "ПРОЕКТ ЭЛЕКТРИКА\\ПРОЕКТ ЭЛЕКТРОСИЛА ОБЩАЯ\\TESLA-ГЕНЕРАТОРЫ СТАРТЕРА",
    ]

    if project == "Корея":
        where.append("группа = ANY(:korea_groups)")
        params["korea_groups"] = korea_groups
    elif project == "Китай":
        where.append("LOWER(группа) LIKE '%china%'")

    sql = f"""
        SELECT дата, склад, товар_id, артикул, наименование,
               остаток, продано, пополнение, цена
        FROM alyans_refresh_v3
        WHERE {' AND '.join(where)}
        ORDER BY дата ASC
    """

    try:
        with engine.connect() as conn:
            res = conn.execute(text(sql), params)
            df = pd.DataFrame(res.fetchall(), columns=res.keys())

        if df.empty:
            return df

        # --- Приведение типов ---
        df["дата"] = pd.to_datetime(df["дата"], errors="coerce")
        for col in ["остаток", "продано", "пополнение", "цена"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        # --- Удаляем дубликаты по дате/складу ---
        df = (
            df.sort_values("дата")
              .drop_duplicates(subset=["дата", "склад"], keep="last")
              .reset_index(drop=True)
        )

        return df

    except Exception:
        logger.exception("[get_product_timeseries] Ошибка получения временного ряда")
        return pd.DataFrame()


# --- Функции подготовки данных ---

def add_canonical_name(df: pd.DataFrame) -> pd.DataFrame:
    """Для каждого (Склад, Артикул, Номенклатура) выбираем каноническое название номенклатуры (мода)."""
    df = df.copy()
    df["Артикул_товар"] = df["Артикул"] + "|" + df["Номенклатура"]

    mode_map = (
        df.groupby(["Склад", "Артикул_товар"])["Номенклатура"]
        .agg(lambda s: s.mode().iat[0] if not s.mode().empty else s.dropna().iloc[0])
    )
    variants_map = (
        df.groupby(["Склад", "Артикул_товар"])["Номенклатура"]
        .agg(lambda s: ", ".join(sorted(set(s.dropna()))))
    )

    idx = df.set_index(["Склад", "Артикул_товар"]).index
    df["Номенклатура_канон"] = idx.map(mode_map.to_dict())
    df["Номенклатура_варианты"] = idx.map(variants_map.to_dict())
    df["Смена_наименования"] = df["Номенклатура"] != df["Номенклатура_канон"]
    return df


def calculate_daily_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Считаем 'Продано' и 'Пришло' по уникальным товарам (Артикул_товар), агрегируем по дате."""
    if df.empty:
        for c in ["Продано", "Пришло", "Цена_изменилась", "Аномалия"]:
            df[c] = pd.Series(dtype=float if c in ["Продано", "Пришло"] else bool)
        return df

    req = ["Склад", "Артикул_товар", "Дата", "Остаток", "Цена"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise ValueError(f"Отсутствуют колонки: {miss}")

    df["Дата_только"] = df["Дата"].dt.normalize()

    df_daily = (
        df.sort_values("Дата")
        .groupby(["Склад", "Артикул_товар", "Дата_только"], as_index=False)
        .agg({
            "Остаток": "first",
            "Цена": "first",
            "Номенклатура": "first",
            "Номенклатура_канон": "first",
            "Номенклатура_варианты": "first"
        })
    )
    df_daily.rename(columns={"Дата_только": "Дата"}, inplace=True)

    g = df_daily.groupby(["Склад", "Артикул_товар"], group_keys=False)
    delta_stock = g["Остаток"].diff()

    df_daily["Продано"] = (-delta_stock.clip(upper=0)).fillna(0)
    df_daily["Пришло"] = (delta_stock.clip(lower=0)).fillna(0)
    df_daily["Цена_изменилась"] = g["Цена"].diff().fillna(0) != 0
    same_ost = delta_stock.fillna(0) == 0
    df_daily["Аномалия"] = ((df_daily["Продано"] > 0) | (df_daily["Пришло"] > 0)) & same_ost

    return df_daily


GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")  # токен с правом push
GITHUB_REPO = "AleksHertz/my-dash"
GITHUB_BRANCH = "main"
ARCHIVE_URL = "https://github.com/AleksHertz/my-dash/raw/refs/heads/main/data/aggregated.zip"
TMP_UPLOAD_PATH = "tmp_uploaded"  # временная папка для загруженных файлов

# --- Вспомогательная функция очистки и унификации DataFrame ---
def unify_and_clean_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    # Унификация колонки Остаток
    if "Количество" in df.columns and "Остаток" not in df.columns:
        df.rename(columns={"Количество": "Остаток"}, inplace=True)

    for col in ["Дата", "Артикул", "Номенклатура", "Остаток", "Склад"]:
        if col not in df.columns:
            df[col] = np.nan

    df["Дата"] = pd.to_datetime(df["Дата"], errors="coerce")
    df["Артикул"] = df["Артикул"].astype(str).str.strip()
    df["Номенклатура"] = df["Номенклатура"].astype(str).str.strip()
    df["Остаток"] = pd.to_numeric(df["Остаток"], errors="coerce")
    df["Цена"] = pd.to_numeric(df["Цена"], errors="coerce") if "Цена" in df.columns else np.nan

    df = df.dropna(subset=["Дата", "Артикул", "Остаток"])
    df = df.drop_duplicates(subset=["Дата", "Артикул", "Склад"], keep="last")
    return df


# --- Загрузка архива с GitHub ---
def load_and_prepare_2025_from_url(url: str = ARCHIVE_URL) -> pd.DataFrame:
    frames = []
    try:
        resp = requests.get(url)
        resp.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
            for name in z.namelist():
                if not name.lower().endswith(".csv"):
                    continue
                with z.open(name) as f:
                    df = pd.read_csv(f)
                    df.columns = [c.strip() for c in df.columns]

                    if "Москва" in name:
                        df["Склад"] = "Москва"
                    elif "Хабаровск" in name:
                        df["Склад"] = "Хабаровск"
                    else:
                        df["Склад"] = "Неизвестно"

                    frames.append(df)
        logging.info(f"[load_and_prepare_2025_from_url] Загружено файлов: {len(frames)}")
        print(f"[load_and_prepare_2025_from_url] Загружено файлов: {len(frames)}")
    except Exception as e:
        logging.error(f"[load_and_prepare_2025_from_url] Ошибка загрузки архива: {e}", exc_info=True)
        print(f"[load_and_prepare_2025_from_url] Ошибка загрузки архива: {e}")

    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    return unify_and_clean_df(df)


# --- Объединённая загрузка данных (архив + новый CSV с GitHub) ---
def load_combined_2025() -> pd.DataFrame:
    print("[load_combined_2025] Вызов функции объединённой загрузки данных")
    logging.info("[load_combined_2025] Вызов функции объединённой загрузки данных")

    try:
        # --- Загружаем архив ---
        df_archive = load_and_prepare_2025_from_url()
        print(f"[load_combined_2025] Архив загружен, строк: {len(df_archive)}")
        logging.info(f"[load_combined_2025] Архив загружен, строк: {len(df_archive)}")

        if df_archive.empty:
            print("[load_combined_2025] Нет данных в архиве")
            logging.warning("[load_combined_2025] Нет данных в архиве")
            return pd.DataFrame()

        # --- Загружаем новые CSV из GitHub ---
        g = Github(GITHUB_TOKEN)
        repo = g.get_repo(GITHUB_REPO)

        all_new_parts = []
        try:
            contents = repo.get_contents("data/new_uploads", ref=GITHUB_BRANCH)
        except Exception as e:
            print(f"[load_combined_2025] Папка data/new_uploads не найдена: {e}")
            logging.warning(f"[load_combined_2025] Папка data/new_uploads не найдена: {e}")
            contents = []

        # Проходим по всем файлам и поддиректориям
        while contents:
            file_content = contents.pop(0)
            if file_content.type == "dir":
                contents.extend(repo.get_contents(file_content.path, ref=GITHUB_BRANCH))
            elif file_content.type == "file" and file_content.path.endswith(".csv"):
                try:
                    csv_data = file_content.decoded_content.decode("utf-8")
                    df_part = pd.read_csv(io.StringIO(csv_data))
                    print(f"[load_combined_2025] Загружен файл {file_content.path}, строк: {len(df_part)}")
                    logging.info(f"[load_combined_2025] Загружен файл {file_content.path}, строк: {len(df_part)}")
                    all_new_parts.append(df_part)
                except Exception as e:
                    print(f"[load_combined_2025] Ошибка чтения {file_content.path}: {e}")
                    logging.error(f"[load_combined_2025] Ошибка чтения {file_content.path}: {e}", exc_info=True)

        df_new = pd.concat(all_new_parts, ignore_index=True) if all_new_parts else pd.DataFrame()
        print(f"[load_combined_2025] Всего новых строк: {len(df_new)}")
        logging.info(f"[load_combined_2025] Всего новых строк: {len(df_new)}")

        # --- Объединяем архив и новые CSV ---
        df = pd.concat([df_archive, df_new], ignore_index=True)
        print(f"[load_combined_2025] После объединения строк: {len(df)}")
        logging.info(f"[load_combined_2025] После объединения строк: {len(df)}")

        # --- Приведение типов ---
        if "Дата" in df.columns:
            df["Дата"] = pd.to_datetime(df["Дата"], errors="coerce")

        # --- Убираем дубликаты ---
        if all(col in df.columns for col in ["Склад", "Артикул", "Дата"]):
            before = len(df)
            df.drop_duplicates(subset=["Склад", "Артикул", "Дата"], inplace=True)
            print(f"[load_combined_2025] Убрано дубликатов: {before - len(df)}")
            logging.info(f"[load_combined_2025] Убрано дубликатов: {before - len(df)}")

        # --- Дополнительная обработка ---
        df = add_canonical_name(df)
        df = calculate_daily_metrics(df)

        print(f"[load_combined_2025] Финальный размер: {len(df)} строк")
        logging.info(f"[load_combined_2025] Финальный размер: {len(df)} строк")
        return df

    except Exception as e:
        print(f"[load_combined_2025] Общая ошибка: {e}")
        logging.error(f"[load_combined_2025] Общая ошибка: {e}", exc_info=True)
        return pd.DataFrame()

# --- Функция загрузки файла в GitHub (универсальная) ---
def github_upload_file(local_path: str, target_path: str, commit_message: str) -> bool:
    try:
        g = Github(GITHUB_TOKEN)
        repo = g.get_repo(GITHUB_REPO)

        with open(local_path, "rb") as f:
            content = f.read()

        try:
            file = repo.get_contents(target_path, ref=GITHUB_BRANCH)
            repo.update_file(
                path=target_path,
                message=commit_message,
                content=content,
                sha=file.sha,
                branch=GITHUB_BRANCH
            )
            logging.info(f"[github_upload_file] Обновлён файл {target_path}")
            print(f"[github_upload_file] Обновлён файл {target_path}")
        except Exception:
            repo.create_file(
                path=target_path,
                message=commit_message,
                content=content,
                branch=GITHUB_BRANCH
            )
            logging.info(f"[github_upload_file] Создан файл {target_path}")
            print(f"[github_upload_file] Создан файл {target_path}")

        return True
    except Exception as e:
        logging.error(f"[github_upload_file] Ошибка загрузки {target_path}: {e}", exc_info=True)
        print(f"[github_upload_file] Ошибка загрузки {target_path}: {e}")
        return False


# --- Инициализация глобальных переменных ---
df_2025 = load_combined_2025()
df_2025_clean = df_2025[~df_2025["Аномалия"]].copy() if not df_2025.empty else pd.DataFrame()

unique_sklads_2025 = sorted(df_2025_clean["Склад"].dropna().unique().tolist()) if not df_2025_clean.empty else []
unique_articles_2025 = sorted(df_2025_clean["Артикул_товар"].dropna().astype(str).unique().tolist()) if not df_2025_clean.empty else []
unique_noms_2025 = sorted(df_2025_clean["Номенклатура_канон"].dropna().unique().tolist()) if not df_2025_clean.empty else []
# --------------------
# DASH APP
# --------------------
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
server = app.server

app.layout = html.Div([
    html.H1("Анализ складских данных"),

    dcc.Tabs(id="tabs", value="main", children=[
        # ===================== Основной анализ =====================
        dcc.Tab(label="Основной анализ", value="main", children=[
            html.Div([
                html.H2("ТОПы по складам"),

                html.Label("Выберите склад:"),
                dcc.Dropdown(
                    id='sklad-filter',
                    options=[{'label': s, 'value': s} for s in unique_sklads],
                    value=unique_sklads,
                    multi=True,
                    placeholder="Выберите один или несколько складов",
                    clearable=True,
                    style={'marginBottom': '20px'}
                ),

                html.Label("Выберите количество позиций для отображения ходовых товаров:"),
                dcc.RadioItems(
                    id='top-n-selector',
                    options=[
                        {'label': 'Топ 100', 'value': 100},
                        {'label': 'Топ 500', 'value': 500},
                        {'label': 'Топ 1000', 'value': 1000},
                    ],
                    value=100,
                    labelStyle={'display': 'inline-block', 'marginRight': '15px'},
                    style={'marginBottom': '20px'}
                ),

                html.H3("Топ самых ходовых товаров"),
                html.Div(
                    dcc.Graph(id='graph-top-fast'),
                    style={
                        'height': '700px',
                        'overflowY': 'scroll',
                        'border': '1px solid #ddd',
                        'padding': '5px',
                        'marginBottom': '10px',
                        'backgroundColor': 'white'
                    }
                ),
                dbc.Button(
                    "📥 Выгрузить топ ходовых в Excel",
                    id="download-top-fast-btn",
                    color="success",
                    className="mb-4"
                ),

                html.Label("Выберите количество позиций для отображения товаров по пополнениям:"),
                dcc.RadioItems(
                    id='top-n-selector-restock',
                    options=[
                        {'label': 'Топ 100', 'value': 100},
                        {'label': 'Топ 500', 'value': 500},
                        {'label': 'Топ 1000', 'value': 1000},
                    ],
                    value=100,
                    labelStyle={'display': 'inline-block', 'marginRight': '15px'},
                    style={'marginBottom': '20px'}
                ),

                html.H3("Топ товаров по пополнениям"),
                html.Div(
                    dcc.Graph(id='graph-top-restock'),
                    style={
                        'height': '700px',
                        'overflowY': 'scroll',
                        'border': '1px solid #ddd',
                        'padding': '5px',
                        'marginBottom': '10px',
                        'backgroundColor': 'white'
                    }
                ),
                dbc.Button(
                    "📥 Выгрузить топ пополнений в Excel",
                    id="download-top-restock-btn",
                    color="success"
                ),

                dcc.Download(id="download-top-fast"),
                dcc.Download(id="download-top-restock"),
            ], style={'marginBottom': 40}),

            html.Div([
                html.H2("Всплески продаж"),
                html.Div([
                    html.Label("Склад:"),
                    dcc.Dropdown(
                        id='peak-sklad-filter',
                        options=[{'label': s, 'value': s} for s in unique_peak_sklads],
                        multi=False,
                        placeholder="Выберите склад для всплесков",
                        clearable=True,
                    ),
                    html.Label("Артикул:"),
                    dcc.Dropdown(
                        id='peak-article-filter',
                        options=[{'label': a, 'value': a} for a in unique_peak_articles],
                        multi=False,
                        placeholder="Выберите артикул",
                        clearable=True,
                    ),
                    html.Label("Номенклатура:"),
                    dcc.Dropdown(
                        id='peak-nom-filter',
                        options=[],
                        multi=False,
                        placeholder="Выберите номенклатуру",
                        clearable=True,
                        searchable=True,
                        style={'width': '100%'}
                    ),
                    html.Button("📥 Скачать в Excel", id="btn-download-peaks", n_clicks=0),
                    dcc.Download(id="download-peaks-xlsx"),
                ], style={
                    'maxWidth': 450,
                    'marginBottom': 30,
                    'display': 'flex',
                    'flexDirection': 'column',
                    'gap': '10px'
                }),
                dcc.Graph(id='graph-peaks'),
                html.Div([
                    html.P("График отображает:"),
                    html.Ul([
                        html.Li("Продажи (оси слева)"),
                        html.Li("Средняя цена (пунктирная линия, правая ось)"),
                        html.Li("Изменение цены в процентах (штриховая линия, правая ось)"),
                    ]),
                ], style={
                    'maxWidth': 600,
                    'fontStyle': 'italic',
                    'color': 'gray',
                    'marginTop': 10
                }),
            ]),
        ]),

        # ===================== Вкладка 2025 =====================
        dcc.Tab(label="Анализ 2025", value="2025", children=[
            html.Div([
                html.H3("Загрузить новые данные"),
                html.Div([
                    html.Span(
                        f"📅 Данные актуальны на: {get_latest_upload_date()}",
                        id="data-update-date",
                        style={"fontSize": "14px", "color": "#555", "marginLeft": "5px"}
                    )
                ], style={"marginBottom": "15px"}),

                dcc.Upload(
                    id='upload-data',
                    children=html.Div(['Перетащите файл сюда или ', html.A('выберите файл')]),
                    style={
                        'width': '100%',
                        'height': '60px',
                        'lineHeight': '60px',
                        'borderWidth': '1px',
                        'borderStyle': 'dashed',
                        'borderRadius': '5px',
                        'textAlign': 'center',
                        'marginBottom': '20px'
                    },
                    multiple=False
                ),
                dcc.Loading(
                    id="loading-upload",
                    type="circle",
                    children=html.Div(id='upload-status', style={'marginTop': '10px', 'color': 'green'})
                ),

                html.Div([
                    html.Label("Склад:"),
                    dcc.Dropdown(
                        id='sklad-2025-filter',
                        options=[{'label': s, 'value': s} for s in unique_sklads_2025],
                        value=unique_sklads_2025,
                        multi=True,
                        placeholder="Выберите склад",
                        clearable=True,
                        style={'marginBottom': '15px'}
                    ),

                    html.Label("Артикул:"),
                    dcc.Dropdown(
                        id='article-2025-filter',
                        options=[],
                        placeholder="Введите или выберите артикул",
                        searchable=True,
                        clearable=True,
                        style={'marginBottom': '15px'}
                    ),
                    html.Div(id="article-hint", style={"fontSize": "13px", "color": "#888", "marginBottom": "15px"}),

                    html.Label("Номенклатура:"),
                    dcc.Dropdown(
                        id='nom-2025-filter',
                        options=[{'label': n, 'value': n} for n in unique_noms_2025],
                        multi=False,
                        placeholder="Выберите номенклатуру",
                        clearable=True,
                        style={'marginBottom': '15px'}
                    ),

                    html.Label("Месяц:"),
                    dcc.Dropdown(
                        id='month-2025-filter',
                        options=[{'label': m, 'value': i+1} for i, m in enumerate([
                            'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
                        ])],
                        multi=False,
                        placeholder="Выберите месяц",
                        clearable=True,
                        style={'marginBottom': '20px'}
                    ),
                ], style={'maxWidth': 500, 'marginBottom': 30}),

                html.H3("Динамика продаж, пополнений и цены выбранного товара"),
                dcc.Graph(id='graph-2025-line'),

                html.Div([
                    html.Label("Размер ТОПа:"),
                    dcc.RadioItems(
                        id="top-size-selector",
                        options=[
                            {"label": "Топ-50", "value": 50},
                            {"label": "Топ-100", "value": 100},
                            {"label": "Топ-250", "value": 250},
                            {"label": "Топ-500", "value": 500},
                        ],
                        value=100,
                        inline=True
                    ),
                ], style={"marginBottom": "10px"}),

                html.H3(id="top-title", style={"marginTop": "20px"}),

                dash_table.DataTable(
                    id="top-100-table",
                    columns=[
                        {"name": "Артикул", "id": "Артикул"},
                        {"name": "Номенклатура", "id": "Номенклатура"},
                        {"name": "Продано", "id": "Продано"},
                        {"name": "Склад", "id": "Склад"},
                    ],
                    style_table={
                        "overflowX": "auto",
                        "maxHeight": "500px",
                        "overflowY": "scroll",
                        "width": "100%"
                    },
                    style_cell={
                        "textAlign": "left",
                        "padding": "5px",
                        "whiteSpace": "normal",
                        "height": "auto"
                    },
                    style_header={
                        "fontWeight": "bold",
                        "backgroundColor": "#f0f0f0"
                    },
                    page_size=20,
                    row_selectable="single",
                ),

                html.Div([
                    dbc.Button(
                        "📥 Выгрузить в Excel (с учётом фильтров)",
                        id="download-2025-btn",
                        color="primary",
                        className="mt-3"
                    ),
                    dcc.Download(id="download-2025-xlsx"),
                ], style={"marginTop": "20px"})
            ])
        ]),

    # ===================== Вкладка Альянс =====================
        dcc.Tab(label="Альянс", value="alyans", children=[
            html.Div([
                html.H2("Анализ данных Альянс"),

                # Загрузка Excel
                html.H4("Загрузка данных"),
                dcc.Upload(
                    id='upload-alyans-data',
                    children=html.Div(['Перетащите файл сюда или ', html.A('выберите файл')]),
                    style={
                        'width': '100%',
                        'height': '60px',
                        'lineHeight': '60px',
                        'borderWidth': '1px',
                        'borderStyle': 'dashed',
                        'borderRadius': '5px',
                        'textAlign': 'center',
                        'marginBottom': '10px'
                    },
                    multiple=False
                ),
                html.Div(id="upload-alyans-status", style={'marginBottom': '10px', 'color': 'green'}),
                dcc.Interval(id='alyans-progress-interval', interval=500, disabled=True),
                html.Div(id="alyans-progress-output", style={'marginBottom': '20px', 'color': 'blue'}),

                # Фильтры
                html.Label("Проект:"),
                dcc.Dropdown(
                    id="project-filter",
                    options=[
                        {"label": "Все проекты", "value": None},
                        {"label": "Корея", "value": "Корея"},
                        {"label": "Китай", "value": "Китай"},
                    ],
                    multi=False,
                    clearable=True,
                    placeholder="Выберите проект",
                    style={"marginBottom": "15px"}
                ),

                html.Label("Склад:"),
                dcc.Dropdown(
                    id="alyans-sklad",
                    multi=True,
                    options=[{"label": s, "value": s} for s in _ensure_list(get_unique_sklads())],
                    placeholder="Выберите склад",
                    style={"marginBottom": "15px"}
                ),

                html.Label("Группа:"),
                dcc.Dropdown(
                    id="alyans-group",
                    multi=True,
                    options=[{"label": g, "value": g} for g in _ensure_list(get_unique_groups())],
                    placeholder="Выберите группу",
                    style={"marginBottom": "20px"}
                ),

                # График
                html.H3("Динамика остатков и продаж по выбранному товару"),
                dcc.Loading(
                    id="loading-alyans-graph",
                    type="circle",
                    children=dcc.Graph(id="alyans-graph")
                ),

                # Размер ТОПа
                html.Div([
                    html.Label("Размер ТОПа:"),
                    dcc.RadioItems(
                        id="alyans-top-size",
                        options=[
                            {"label": "Топ-100", "value": 100},
                            {"label": "Топ-250", "value": 250},
                            {"label": "Топ-500", "value": 500},
                            {"label": "Топ-1000", "value": 1000},
                            {"label": "Топ-2000", "value": 2000},
                        ],
                        value=100,
                        inline=True
                    ),
                ], style={"marginBottom": "10px"}),

                html.H3(id="alyans-top-title", style={"marginTop": "20px"}),

                # Таблица
                dcc.Loading(
                    id="loading-alyans-table",
                    type="circle",
                    children=dash_table.DataTable(
                        id="alyans-table",
                        columns=[
                            {"name": "Артикул", "id": "Артикул"},
                            {"name": "Наименование", "id": "Наименование"},
                            {"name": "Продано", "id": "Продано"},
                            {"name": "Склад", "id": "Склад"},
                        ],
                        style_table={
                            "overflowX": "auto",
                            "maxHeight": "500px",
                            "overflowY": "scroll",
                            "width": "100%"
                        },
                        style_cell={
                            "textAlign": "left",
                            "padding": "5px",
                            "whiteSpace": "normal",
                            "height": "auto"
                        },
                        style_header={
                            "fontWeight": "bold",
                            "backgroundColor": "#f0f0f0"
                        },
                        page_size=20,
                        row_selectable="single",
                    )
                ),

                # Кнопка выгрузки
                html.Div([
                    dbc.Button(
                        "📥 Выгрузить в Excel (с учётом фильтров)",
                        id="download-alyans-btn",
                        color="primary",
                        className="mt-3"
                    ),
                    dcc.Download(id="download-alyans-xlsx"),
                ], style={"marginTop": "20px"})
            ])
        ]),
    ])
])

# --------------------
# КОЛБЭКИ
# --------------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# -------------------- Колбэк загрузки --------------------
@app.callback(
    Output("upload-alyans-status", "children"),
    Output("alyans-progress-interval", "disabled"),
    Input("upload-alyans-data", "contents"),
    State("upload-alyans-data", "filename"),
    prevent_initial_call=True
)
def handle_alyans_upload(content, filename):
    """Обрабатывает Excel и готовит данные для вкладки Альянс"""
    global upload_alyans_df
    if not content:
        return "❌ Файл не загружен", True

    last_date = get_last_date_in_db()
    file_date = extract_date_from_filename(filename)
    if not file_date:
        return "⚠️ Не удалось определить дату из имени файла", True
    if last_date and file_date <= last_date:
        return f"⚠️ Файл {filename} содержит старые данные (до {last_date})", True

    # Декодируем Excel
    content_type, content_string = content.split(',')
    decoded = base64.b64decode(content_string)
    try:
        df = pd.read_excel(io.BytesIO(decoded), dtype=str)
    except Exception as e:
        return f"❌ Ошибка чтения Excel: {e}", True

    df = clean_articles(df)

    sklad_cols = [c for c in df.columns if c.strip().startswith("Остаток")]
    if not sklad_cols:
        return f"⚠️ В файле нет колонок Остаток...", True

    result = []
    for col in sklad_cols:
        temp = df.copy()
        temp["остаток"] = pd.to_numeric(temp[col], errors="coerce").fillna(0).astype(int)
        temp["склад"] = col.replace("Остаток", "").strip()
        temp["дата"] = file_date
        temp["id"] = [uuid.uuid4() for _ in range(len(temp))]
        result.append(temp)

    df_new = pd.concat(result, ignore_index=True)

    # Загружаем остатки предыдущей даты
    if last_date:
        df_prev = pd.read_sql(
            text(f"SELECT товар_id, склад, остаток FROM {TABLE_NAME} WHERE дата = :d"),
            engine,
            params={"d": last_date}
        )
    else:
        df_prev = pd.DataFrame(columns=["товар_id", "склад", "остаток"])

    df_ready = calc_changes(df_new, df_prev)
    upload_alyans_df = df_ready

    return f"🔄 Начинается загрузка {len(df_ready):,} строк за {file_date}...", False

# -------------------- Колбэк прогресса для Альянса --------------------
@app.callback(
    Output("alyans-progress-output", "children"),
    Output("upload-alyans-status", "children", allow_duplicate=True),
    Input("alyans-progress-interval", "n_intervals"),
    prevent_initial_call=True
)
def update_alyans_progress(n):
    """Пошаговая запись данных в базу для Альянса"""
    global upload_alyans_df
    if upload_alyans_df is None or len(upload_alyans_df) == 0:
        return "⚠️ Нет данных для загрузки", "⚠️ Загрузка отменена"

    try:
        if n == 0:
            # Создаем генератор для пошаговой записи
            update_alyans_progress.generator = save_to_db(upload_alyans_df)

        percent = next(update_alyans_progress.generator)
        return f"⏳ Загрузка: {percent}%", f"Загрузка выполняется..."

    except StopIteration:
        upload_alyans_df = None
        last_date = get_last_date_in_db()
        return f"✅ Загрузка завершена на 100%", f"✅ Последняя дата в БД: {last_date}"

    except Exception as e:
        upload_alyans_df = None
        return f"❌ Ошибка: {e}", "❌ Ошибка при загрузке данных"


# -------------------- Колбэк для фильтров --------------------
# ---- Dash callbacks: фильтры (load_filters) ----
@app.callback(
    Output("alyans-filters", "children"),
    Input("tabs", "active_tab")
)
def load_filters(active_tab):
    if active_tab != "alyans":
        return []

    sklads = get_unique_sklads()
    groups = get_unique_groups()

    return [
        html.Label("Склад"),
        dcc.Dropdown(id="alyans-sklad", options=[{"label": s, "value": s} for s in sklads], multi=True),
        html.Label("Группа"),
        dcc.Dropdown(id="alyans-group", options=[{"label": g, "value": g} for g in groups], multi=True),
        html.Label("ТОП товаров"),
        dcc.Input(id="alyans-top-size", type="number", value=50, min=10, max=200)
    ]


# ---------- Колбэки ----------
# ---------- Колбэки ----------

@app.callback(
    Output("alyans-table", "data"),
    Output("alyans-table", "selected_rows"),
    Output("alyans-top-title", "children"),
    Input("alyans-sklad", "value"),
    Input("alyans-group", "value"),
    Input("project-filter", "value"),
    Input("alyans-top-size", "value"),
)
def update_alyans_table(selected_sklads, selected_groups, selected_project, top_n):
    sklads = _ensure_list(selected_sklads)
    groups = _ensure_list(selected_groups)
    top_n = int(top_n or 100)

    if not sklads and not groups and not selected_project:
        return [], [], "Выберите склад, группу или проект"

    df = get_top_products(
        top_n=top_n,
        sklads=sklads,
        groups=groups,
        project=selected_project
    )

    if df.empty:
        return [], [], f"ТОП-{top_n}: нет данных"

    # Добавляем отсутствующие колонки
    for col in ["артикул", "наименование", "всего_продано", "всего_пополнено", "склад", "товар_id"]:
        if col not in df.columns:
            df[col] = None

    # Формируем таблицу с отображением нужных полей
    df = df.rename(columns={
        "артикул": "Артикул",
        "наименование": "Наименование",
        "всего_продано": "Продано",
        "всего_пополнено": "Пополнено",
        "склад": "Склад",
        "товар_id": "Товар ID"
    })

    # Сортируем и ограничиваем вывод (если данных больше, чем нужно)
    df = df.sort_values("Продано", ascending=False).head(top_n)
    df = df[["Артикул", "Наименование", "Продано", "Пополнено", "Склад", "Товар ID"]]

    # Корректный заголовок с активными фильтрами
    title_parts = [f"ТОП-{top_n} товаров"]
    if sklads:
        title_parts.append(f"по складам: {', '.join(sklads)}")
    if groups:
        title_parts.append(f"по группам: {', '.join(groups)}")
    if selected_project:
        title_parts.append(f"проект: {selected_project}")

    title = " | ".join(title_parts)

    return df.to_dict("records"), [], title



@app.callback(
    Output("alyans-graph", "figure"),
    Input("alyans-table", "data"),
    Input("alyans-table", "selected_rows")
)
def update_alyans_graph(table_data, selected_rows):
    if not table_data or not selected_rows:
        return go.Figure(layout=go.Layout(title="Выберите товар из таблицы"))

    sel = table_data[selected_rows[0]]
    tovar_id = sel.get("Товар ID") or sel.get("Артикул")
    skl = sel.get("Склад")

    if not tovar_id:
        return go.Figure(layout=go.Layout(title="Нет идентификатора товара"))

    df = get_product_timeseries(
        tovar_id=tovar_id,
        sklads=[skl] if skl and skl != "Суммарно" else None
    )
    if df.empty:
        return go.Figure(layout=go.Layout(title=f"Нет данных по товару {tovar_id}"))

    df = df.sort_values("дата")
    df["дата"] = pd.to_datetime(df["дата"])
    for col in ["остаток", "продано", "пополнение", "цена"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["price_prev"] = df["цена"].shift(1)
    df["price_change"] = ""
    df.loc[df["цена"] > df["price_prev"], "price_change"] = "⬆️"
    df.loc[df["цена"] < df["price_prev"], "price_change"] = "⬇️"

    df["hover_text"] = df.apply(
        lambda r: f"Склад: {r.get('склад', '—')}<br>"
                  f"Дата: {r['дата'].date()}<br>"
                  f"Продано: {int(r['продано'])}<br>"
                  f"Пополнено: {int(r['пополнение'])}<br>"
                  f"Остаток: {int(r['остаток'])}<br>"
                  f"Цена: {r['цена']:.2f} {r['price_change']}",
        axis=1
    )

    fig = go.Figure()

    # --- Много складов ---
    if df["склад"].nunique() > 1:
        for skl_name, sub in df.groupby("склад"):
            sub_agg = sub.groupby("дата", as_index=False).agg({
                "продано": "sum",
                "пополнение": "sum",
                "остаток": "mean",
                "цена": "mean"
            })
            sub_agg["price_prev"] = sub_agg["цена"].shift(1)
            sub_agg["price_change"] = ""
            sub_agg.loc[sub_agg["цена"] > sub_agg["price_prev"], "price_change"] = "⬆️"
            sub_agg.loc[sub_agg["цена"] < sub_agg["price_prev"], "price_change"] = "⬇️"
            sub_agg["hover_text"] = sub_agg.apply(
                lambda r: f"Склад: {skl_name}<br>"
                          f"Дата: {r['дата'].date()}<br>"
                          f"Продано: {int(r['продано'])}<br>"
                          f"Пополнено: {int(r['пополнение'])}<br>"
                          f"Остаток: {int(r['остаток'])}<br>"
                          f"Цена: {r['цена']:.2f} {r['price_change']}",
                axis=1
            )

            fig.add_trace(go.Bar(
                x=sub_agg["дата"], y=sub_agg["продано"],
                name=f"Продано ({skl_name})", opacity=0.7,
                hovertext=sub_agg["hover_text"], hoverinfo="text"
            ))
            fig.add_trace(go.Bar(
                x=sub_agg["дата"], y=sub_agg["пополнение"],
                name=f"Пополнено ({skl_name})", opacity=0.7,
                hovertext=sub_agg["hover_text"], hoverinfo="text"
            ))
            fig.add_trace(go.Scatter(
                x=sub_agg["дата"], y=sub_agg["остаток"],
                mode="lines+markers", name=f"Остаток ({skl_name})",
                yaxis="y2", line=dict(width=2),
                hovertext=sub_agg["hover_text"], hoverinfo="text"
            ))

            price_change_points = sub_agg[sub_agg["price_change"] != ""]
            if not price_change_points.empty:
                fig.add_trace(go.Scatter(
                    x=price_change_points["дата"], y=price_change_points["остаток"],
                    mode="markers",
                    marker=dict(
                        size=12,
                        color=price_change_points["price_change"].map({"⬆️":"green","⬇️":"red"}),
                        symbol=price_change_points["price_change"].map({"⬆️":"triangle-up","⬇️":"triangle-down"})
                    ),
                    name=f"Изм. цены ({skl_name})",
                    hovertext=price_change_points["hover_text"],
                    hoverinfo="text",
                    yaxis="y2"
                ))
    # --- Один склад ---
    else:
        agg = df.groupby("дата", as_index=False).agg({
            "продано": "sum",
            "пополнение": "sum",
            "остаток": "mean",
            "цена": "mean"
        })
        agg["price_prev"] = agg["цена"].shift(1)
        agg["price_change"] = ""
        agg.loc[agg["цена"] > agg["price_prev"], "price_change"] = "⬆️"
        agg.loc[agg["цена"] < agg["price_prev"], "price_change"] = "⬇️"
        agg["hover_text"] = agg.apply(
            lambda r: f"Склад: {skl or '—'}<br>"
                      f"Дата: {r['дата'].date()}<br>"
                      f"Продано: {int(r['продано'])}<br>"
                      f"Пополнено: {int(r['пополнение'])}<br>"
                      f"Остаток: {int(r['остаток'])}<br>"
                      f"Цена: {r['цена']:.2f} {r['price_change']}",
            axis=1
        )

        fig.add_trace(go.Bar(x=agg["дата"], y=agg["продано"], name="Продано",
                             marker_color="steelblue", hovertext=agg["hover_text"], hoverinfo="text"))
        fig.add_trace(go.Bar(x=agg["дата"], y=agg["пополнение"], name="Пополнено",
                             marker_color="lightgreen", hovertext=agg["hover_text"], hoverinfo="text"))
        fig.add_trace(go.Scatter(
            x=agg["дата"], y=agg["остаток"], mode="lines+markers",
            name="Остаток", yaxis="y2", line=dict(color="firebrick", width=2),
            hovertext=agg["hover_text"], hoverinfo="text"
        ))

        price_change_points = agg[agg["price_change"] != ""]
        if not price_change_points.empty:
            fig.add_trace(go.Scatter(
                x=price_change_points["дата"], y=price_change_points["остаток"],
                mode="markers",
                marker=dict(
                    size=12,
                    color=price_change_points["price_change"].map({"⬆️":"green","⬇️":"red"}),
                    symbol=price_change_points["price_change"].map({"⬆️":"triangle-up","⬇️":"triangle-down"})
                ),
                name="Изм. цены",
                hovertext=price_change_points["hover_text"],
                hoverinfo="text",
                yaxis="y2"
            ))

     # --- Аннотация внизу ---
    total_sold = int(df["продано"].sum())
    total_refilled = int(df["пополнение"].sum())

    # 🧮 Берём последний день и суммируем остатки по всем складам на эту дату
    last_date = df["дата"].max()
    last_stock = int(df.loc[df["дата"] == last_date, "остаток"].sum())

    fig.add_annotation(
        text=(
            f"💰 Всего продано: {total_sold:,} | "
            f"🔄 Пополнено: {total_refilled:,} | "
            f"📦 Остаток (на {last_date.date()}): {last_stock:,}"
        ),
        showarrow=False,
        xref="paper", yref="paper",
        x=0.5, y=-0.22,
        font=dict(size=13, color="gray")
    )

    fig.update_layout(
        title=f"📦 Динамика продаж и остатков — {sel.get('Наименование', '')} ({tovar_id})",
        xaxis_title="Дата",
        yaxis_title="Продано / Пополнено",
        yaxis2=dict(title="Остаток", overlaying="y", side="right", showgrid=False),
        barmode="group",
        template="plotly_white",
        height=620,
        legend=dict(orientation="h", yanchor="bottom", y=-0.1),
        margin=dict(l=60, r=60, t=70, b=130)  # увеличенный нижний отступ
    )

    return fig


@app.callback(
    Output("download-alyans-xlsx", "data"),
    Input("download-alyans-btn", "n_clicks"),
    State("alyans-table", "data"),
    State("alyans-table", "selected_rows"),
    prevent_initial_call=True
)
def download_alyans_xlsx(n_clicks, table_data, selected_rows):
    if not table_data or not selected_rows:
        return None

    sel = table_data[selected_rows[0]]
    tovar_id = sel.get("Товар ID") or sel.get("Артикул")
    skl = sel.get("Склад")

    df = get_product_timeseries(
        tovar_id=tovar_id,
        sklads=[skl] if skl and skl != "Суммарно" else None
    )
    if df.empty:
        return None

    df = df.sort_values("дата")
    df["дата"] = pd.to_datetime(df["дата"])
    for col in ["остаток", "продано", "пополнение", "цена"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # --- Создаём Excel в памяти ---
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Свод"

    # --- Общая агрегированная таблица ---
    agg = df.groupby("дата", as_index=False).agg({
        "продано": "sum",
        "пополнение": "sum",
        "остаток": "sum",
        "цена": "mean"
    })
    agg.rename(columns={"цена": "Средняя цена, ₽"}, inplace=True)

    # Запись данных в сводную вкладку
    for r in dataframe_to_rows(agg, index=False, header=True):
        ws_summary.append(r)

    # --- Добавляем вкладки по складам ---
    for skl_name, sub in df.groupby("склад"):
        ws = wb.create_sheet(title=str(skl_name)[:31])  # Excel ограничивает длину названия
        sub_out = sub[["дата", "продано", "пополнение", "остаток", "цена"]].copy()
        sub_out.rename(columns={"цена": "Цена, ₽"}, inplace=True)
        for r in dataframe_to_rows(sub_out, index=False, header=True):
            ws.append(r)

        # форматирование колонок
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    # --- Форматируем свод ---
    for col in ws_summary.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws_summary.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.freeze_panes = "A2"

    # --- Метаданные и стиль ---
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    # --- Сохраняем в буфер ---
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"Отчёт_{tovar_id}.xlsx"
    return dcc.send_bytes(bio.getvalue(), filename)


# --- Утилиты ---
def _to_list(x):
    """Нормализуем значение из дропдауна: str -> [str], None -> []"""
    if x is None:
        return []
    if isinstance(x, (list, tuple, set)):
        return list(x)
    return [x]

# ===================== Функции =====================

def get_item_line(df, article=None, nom=None, sklad_filter=None):
    dff = df.copy()
    sklads = _to_list(sklad_filter)
    if sklads:
        dff = dff[dff["Склад"].isin(sklads)]
    if article:
        dff = dff[dff["Артикул_товар"].astype(str) == str(article)]
    if nom:
        dff = dff[dff["Номенклатура_канон"] == nom]
    dff = dff.sort_values("Дата")

    keep = ["Дата", "Склад", "Артикул_товар", "Номенклатура_канон", "Остаток",
            "Продано", "Пришло", "Цена", "Цена_изменилась", "Аномалия"]
    return dff[keep]


def normalize_dataframe(df: pd.DataFrame, filename: str) -> pd.DataFrame:
    """Очистка и нормализация данных"""
    rows_before = len(df)

    # Приведение дат
    df["Дата"] = pd.to_datetime(df["Дата"], errors="coerce")
    bad_dates = df["Дата"].isna().sum()
    if bad_dates > 0:
        logging.warning(f"[{filename}] {bad_dates} строк удалено из-за некорректной даты")
        df = df.dropna(subset=["Дата"])

    # Заполнение пустых артикулов
    df["Артикул"] = df["Артикул"].fillna("").astype(str)
    missing_articles = (df["Артикул"] == "").sum()
    if missing_articles > 0:
        logging.warning(f"[{filename}] {missing_articles} строк с пустым артикулом")
        df.loc[df["Артикул"] == "", "Артикул"] = [
            f"UNKNOWN_{i}" for i in range(missing_articles)
        ]

    rows_after = len(df)
    logging.info(f"[{filename}] нормализация завершена: {rows_before} → {rows_after} строк")
    return df


# --- Колбэк загрузки файла через Dash ---
@app.callback(
    Output("upload-status", "children"),
    Input("upload-data", "contents"),
    State("upload-data", "filename"),
    prevent_initial_call=True
)
def upload_2025_file(contents, filename):
    if contents is None:
        raise dash.exceptions.PreventUpdate

    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        os.makedirs(TMP_UPLOAD_PATH, exist_ok=True)
        tmp_path = os.path.join(TMP_UPLOAD_PATH, filename)
        with open(tmp_path, "wb") as f:
            f.write(decoded)

        # читаем Excel → DataFrame
        df_new = read_excel_file(tmp_path, sklad_name="auto")
        if df_new is None or df_new.empty:
            return f"Файл {filename} пуст или некорректен"

        # дата из A2
        try:
            file_date = pd.to_datetime(df_new.iloc[1, 0], dayfirst=True, errors="coerce")
            if pd.isna(file_date):
                file_date = datetime.now()
        except Exception:
            file_date = datetime.now()

        # сохраняем в CSV
        date_str = file_date.strftime("%Y-%m-%d")
        time_str = datetime.now().strftime("%H-%M")
        folder_path = os.path.join(TMP_UPLOAD_PATH, "new_uploads")
        os.makedirs(folder_path, exist_ok=True)
        csv_path = os.path.join(folder_path, f"new_uploads_{date_str}_{time_str}.csv")
        df_new.to_csv(csv_path, index=False, encoding="utf-8-sig")

        # пушим на GitHub
        upload_new_csv_to_github(csv_path)

        # сообщение пользователю
        return f"Файл {filename} загружен и отправлен в GitHub. Railway перезапустит приложение — обновите страницу через минуту."

    except Exception as e:
        logging.error(f"[upload_2025_file] Ошибка: {e}", exc_info=True)
        return f"Ошибка при загрузке файла {filename}: {e}"


# --- Фоновая функция для запуска github_upload_file ---
def upload_new_csv_to_github(csv_path: str):
    logging.info(f"[upload_new_csv_to_github] Старт загрузки {csv_path}")
    print(f"[upload_new_csv_to_github] Старт загрузки {csv_path}")

    target_filename = os.path.basename(csv_path)
    relative_path = f"data/new_uploads/{target_filename}"
    commit_msg = f"Добавление новых данных: {target_filename}"

    github_upload_file(csv_path, relative_path, commit_msg)
# ------------------- Колбэк графика -------------------
@app.callback(
    Output("graph-2025-line", "figure"),
    Input("article-2025-filter", "value"),
    Input("nom-2025-filter", "value"),
    Input("sklad-2025-filter", "value"),
    Input("month-2025-filter", "value")
)
def update_line_graph(selected_article, selected_nom, selected_sklads, selected_month):
    # Если товар не выбран
    if not selected_article or not selected_nom:
        return go.Figure(
            layout=go.Layout(
                title="Выберите товар из таблицы ТОП-100 для отображения графика",
                xaxis_title="Дата",
                yaxis_title="Остаток"
            )
        )

    dff = df_2025_clean.copy()

    # --- Фильтр по складам ---
    if selected_sklads:
        dff = dff[dff["Склад"].isin(_to_list(selected_sklads))]

    # --- Гибкий фильтр по артикулу (игнорируем тире, регистр) ---
    dff = dff[dff["Артикул_товар"].astype(str)
              .str.replace("-", "")
              .str.contains(str(selected_article).replace("-", ""), case=False, na=False)]

    # --- Фильтр по номенклатуре ---
    dff = dff[dff["Номенклатура_канон"] == selected_nom]

    # --- Фильтр по месяцу ---
    if selected_month:
        dff = dff[dff["Дата"].dt.month == selected_month]

    if dff.empty:
        return go.Figure(
            layout=go.Layout(
                title="Нет данных для выбранного товара",
                xaxis_title="Дата",
                yaxis_title="Остаток"
            )
        )

    # --- Построение графика ---
    fig = go.Figure()

    for sklad in dff["Склад"].unique():
        df_s = dff[dff["Склад"] == sklad].sort_values("Дата").copy()

        df_s["Продано_fix"] = (df_s["Остаток"].shift(1) - df_s["Остаток"]).clip(lower=0).fillna(0)
        df_s["Пополнено_fix"] = (df_s["Остаток"] - df_s["Остаток"].shift(1)).clip(lower=0).fillna(0)
        df_s["Среднее_Продано"] = df_s["Продано_fix"].rolling(7, min_periods=1).mean()

        df_s["Всплеск"] = df_s["Продано_fix"] > 1.5 * df_s["Среднее_Продано"]
        df_s["Цена_изменилась"] = df_s["Цена"].diff().fillna(0) != 0

        df_s["Цвет"] = df_s.apply(
            lambda row: "purple" if row["Всплеск"] and row["Цена_изменилась"]
            else "red" if row["Всплеск"]
            else "green" if row["Цена_изменилась"]
            else "blue", axis=1
        )
        df_s["Размер"] = df_s["Всплеск"].apply(lambda x: 10 if x else 5)

        fig.add_trace(go.Scatter(
            x=df_s["Дата"],
            y=df_s["Остаток"],
            mode="lines+markers",
            name=str(sklad),
            marker=dict(size=df_s["Размер"], color=df_s["Цвет"]),
            text=[sklad] * len(df_s),
            customdata=df_s[[
                "Продано_fix", "Пополнено_fix", "Цена",
                "Артикул_товар", "Номенклатура_канон",
                "Всплеск", "Цена_изменилась"
            ]].values,
            hovertemplate=(
                "<b>Склад:</b> %{text}<br>"
                "<b>Дата:</b> %{x|%d-%m-%Y}<br>"
                "<b>Остаток:</b> %{y}<br>"
                "<b>Продано:</b> %{customdata[0]}<br>"
                "<b>Пополнено:</b> %{customdata[1]}<br>"
                "<b>Цена:</b> %{customdata[2]}<br>"
                "<b>Артикул:</b> %{customdata[3]}<br>"
                "<b>Номенклатура:</b> %{customdata[4]}<br>"
                "<b>Всплеск:</b> %{customdata[5]}<br>"
                "<b>Изм. цены:</b> %{customdata[6]}<br><extra></extra>"
            ),
            showlegend=False
        ))

    # --- Легенда ---
    legend_colors = {
        "Всплеск": "red",
        "Изменение цены": "green",
        "Всплеск + Изм. цены": "purple",
        "Обычный день": "blue"
    }
    for label, color in legend_colors.items():
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=8, color=color), name=label
        ))

    fig.update_layout(
        title="Динамика остатков, продаж и цен (2025)",
        xaxis_title="Дата",
        yaxis_title="Остаток",
        hovermode="closest",
        legend=dict(orientation="h", y=-0.2)
    )

    return fig

# ------------------- Таблица топ-100 -------------------
# ------------------- Таблица ТОП-N -------------------
@app.callback(
    Output("top-100-table", "data"),
    Output("top-100-table", "selected_rows"),  # сохраняем/сбрасываем выбор
    Output("top-title", "children"),           # <-- динамический заголовок
    Input("sklad-2025-filter", "value"),
    Input("top-size-selector", "value"),       # <-- выбор размера ТОПа
    State("top-100-table", "data"),
    State("top-100-table", "selected_rows"),
)
def update_top_table(selected_sklads, top_n, prev_data, prev_selected):
    dff = df_2025_clean.copy()

    # Фильтр по складам
    if selected_sklads:
        dff = dff[dff["Склад"].isin(_to_list(selected_sklads))]

    if dff.empty:
        return [], [], f"ТОП-{top_n} товаров по продажам (2025)"

    # Группировка по артикулу + номенклатуре + складу
    top_df = (
        dff.groupby(["Артикул_товар", "Номенклатура_канон", "Склад"], as_index=False)
           .agg({"Продано": "sum"})
           .sort_values("Продано", ascending=False)
           .head(top_n)
    )

    # Переименовываем для таблицы
    top_df = top_df.rename(
        columns={
            "Артикул_товар": "Артикул",
            "Номенклатура_канон": "Номенклатура",
        }
    )

    records = top_df.to_dict("records")

    # --- Попытка сохранить выбор ---
    if prev_selected and prev_data:
        try:
            old_row = prev_data[prev_selected[0]]
            # Ищем товар по Артикулу + Номенклатуре
            for idx, row in enumerate(records):
                if (
                    row["Артикул"] == old_row["Артикул"]
                    and row["Номенклатура"] == old_row["Номенклатура"]
                ):
                    return records, [idx], f"ТОП-{top_n} товаров по продажам (2025)"
        except Exception:
            pass

    # Если не нашли → сбрасываем выбор
    return records, [], f"ТОП-{top_n} товаров по продажам (2025)"

# --- нормализация артикула (оставляем) ---
def normalize_article(article):
    """Удаляет все нецифровые символы"""
    if not isinstance(article, str):
        return ""
    return re.sub(r"\D", "", article)





# подготовка словарей — выполнить один раз при старте (после того, как unique_articles_2025 определён)
ALL_ARTICLES_2025 = [str(a) for a in unique_articles_2025 if a is not None]
# карта article -> normalized digits
ALL_ARTICLES_NORM = {a: re.sub(r"\D", "", a) for a in ALL_ARTICLES_2025}
# обратная: normalized -> list(articles)
NORM_TO_ARTS = defaultdict(list)
for a, norm in ALL_ARTICLES_NORM.items():
    NORM_TO_ARTS[norm].append(a)

# -------------------
# callback: динамические options для article-2025-filter
# -------------------
# --- Автоподбор артикулов по введённому тексту ---
@app.callback(
    Output("article-2025-filter", "options"),
    Input("article-2025-filter", "search_value")
)
def update_article_options(search_value):
    """
    Подбирает варианты артикула при вводе.
    Учитывает дефисы и регистр, поддерживает вставку без символов.
    """
    if not search_value:
        # при пустом вводе — первые 50
        return [{"label": a, "value": a} for a in ALL_ARTICLES_2025[:50]]

    # нормализуем ввод
    norm_search = normalize_article(search_value)

    # подбираем совпадения по цифрам (без дефисов)
    matches = [
        a for a, norm in ALL_ARTICLES_NORM.items()
        if norm_search in norm or norm in norm_search
    ]

    # если прямых нет — пробуем fuzzy (через difflib)
    if not matches:
        from difflib import get_close_matches
        all_norm = list(ALL_ARTICLES_NORM.values())
        close_norms = get_close_matches(norm_search, all_norm, n=50, cutoff=0.6)
        matches = [
            a for a, norm in ALL_ARTICLES_NORM.items()
            if norm in close_norms
        ]

    # ограничиваем до 50
    matches = list(dict.fromkeys(matches))[:50]

    if not matches:
        return [{"label": f"❌ Нет совпадений для '{search_value}'", "value": None}]

    return [{"label": a, "value": a} for a in matches]



# -------------------
# Объединённый колбэк: синхронизация артикула / номенклатуры / выбор из таблицы
# (оставим примерно ваш вариант, но безопасно)
# -------------------
@app.callback(
    Output("article-2025-filter", "value"),
    Output("nom-2025-filter", "value"),
    Output("month-2025-filter", "value"),
    Input("top-100-table", "selected_rows"),
    Input("top-100-table", "data"),
    Input("article-2025-filter", "value"),
    Input("nom-2025-filter", "value"),
    prevent_initial_call=True,
)
def sync_article_and_nom(selected_rows, table_data, article_value, nom_value):
    ctx = dash.callback_context
    if not ctx.triggered:
        return dash.no_update, dash.no_update, dash.no_update
    trigger = ctx.triggered[0]["prop_id"]

    # 1) выбор из таблицы
    if trigger.startswith("top-100-table.selected_rows"):
        if selected_rows and table_data:
            try:
                row = table_data[selected_rows[0]]
                art = row.get("Артикул")
                nom = row.get("Номенклатура")
                return art, nom, None
            except Exception:
                return dash.no_update, dash.no_update, dash.no_update
        return dash.no_update, dash.no_update, dash.no_update

    # 2) изменение артикула — подставляем номенклатуру (если есть)
    if trigger.startswith("article-2025-filter.value"):
        if not article_value:
            return article_value, dash.no_update, dash.no_update
        try:
            mask = df_2025_clean["Артикул_товар"].astype(str).str.replace(r"\D", "", regex=True) == re.sub(r"\D", "", str(article_value))
            df_row = df_2025_clean[mask]
            if not df_row.empty:
                nom_match = df_row["Номенклатура_канон"].iloc[0]
                return article_value, nom_match, dash.no_update
        except Exception:
            return article_value, dash.no_update, dash.no_update
        return article_value, dash.no_update, dash.no_update

    # 3) изменение номенклатуры — подставляем артикул (если есть)
    if trigger.startswith("nom-2025-filter.value"):
        if not nom_value:
            return dash.no_update, nom_value, dash.no_update
        try:
            mask = df_2025_clean["Номенклатура_канон"] == nom_value
            df_row = df_2025_clean[mask]
            if not df_row.empty:
                article_match = df_row["Артикул_товар"].astype(str).iloc[0]
                return article_match, nom_value, dash.no_update
        except Exception:
            return dash.no_update, nom_value, dash.no_update
        return dash.no_update, nom_value, dash.no_update

    return dash.no_update, dash.no_update, dash.no_update
# --- Выгрузка топ-ходовых ---

def format_excel(dff, writer, sheet_name):
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]

    # Форматы
    money_fmt = workbook.add_format({'num_format': '#,##0.00 ₽'})
    integer_fmt = workbook.add_format({'num_format': '#,##0'})
    percent_fmt = workbook.add_format({'num_format': '0.00%'})

    # Автоширина колонок и форматы
    for i, col in enumerate(dff.columns):
        max_len = max(
            dff[col].astype(str).map(len).max(),
            len(col)
        ) + 2

        fmt = None
        if col in ['Цена_в_начале', 'Цена_в_конце', 'Средняя_цена', 'Мин_цена', 'Макс_цена']:
            fmt = money_fmt
        elif col in ['Продано', 'Всего_пополнено', 'Средний_остаток']:
            fmt = integer_fmt
        elif col == 'Изменение_цены_%':
            fmt = percent_fmt
        elif col == 'Оборачиваемость':
            fmt = integer_fmt

        worksheet.set_column(i, i, max_len, fmt)

# --- Callback для топ-ходовых ---
@app.callback(
    Output("download-top-fast", "data"),
    Input("download-top-fast-btn", "n_clicks"),
    State("sklad-filter", "value"),
    State("top-n-selector", "value"),
    prevent_initial_call=True
)
def export_top_fast_to_excel(n_clicks, selected_sklads, top_n):
    if df_fast.empty or not selected_sklads:
        return None

    dff = df_fast[df_fast['Склад'].isin(selected_sklads)]
    dff = dff.sort_values('Всего_продано', ascending=False).head(top_n)

    # Новые расчёты
    for col in ['Средняя_цена', 'Мин_цена', 'Макс_цена']:
        if col in dff.columns:
            dff[col] = dff[col].round(2)

    if 'Цена_в_начале' in dff.columns and 'Цена_в_конце' in dff.columns:
        dff['Изменение_цены_%'] = (
            (dff['Цена_в_конце'] - dff['Цена_в_начале']) / dff['Цена_в_начале']
        ).round(4)

    if 'Средний_остаток' in dff.columns:
        dff['Оборачиваемость'] = (dff['Всего_продано'] / dff['Средний_остаток']).round(2)

    # Переименование колонок
    dff = dff.rename(columns={
        'Дней_продаж': 'Количество раз продаж',
        'Дней_в_наличии': 'Количество раз в наличии'
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        dff.to_excel(writer, index=False, sheet_name="Топ_ходовые")
        format_excel(dff, writer, sheet_name="Топ_ходовые")
    output.seek(0)

    return dcc.send_bytes(output.getvalue(), filename=f"топ_{top_n}_ходовые.xlsx")

# --- Callback для топ-пополнений ---
@app.callback(
    Output("download-top-restock", "data"),
    Input("download-top-restock-btn", "n_clicks"),
    State("sklad-filter", "value"),
    State("top-n-selector-restock", "value"),
    prevent_initial_call=True
)
def export_top_restock_to_excel(n_clicks, selected_sklads, top_n):
    if df_restock.empty or not selected_sklads:
        return None

    dff = df_restock[df_restock['Склад'].isin(selected_sklads)]
    dff = dff.sort_values('Всего_пополнено', ascending=False).head(top_n)

    # Новые расчёты
    for col in ['Средняя_цена', 'Мин_цена', 'Макс_цена']:
        if col in dff.columns:
            dff[col] = dff[col].round(2)

    if 'Цена_в_начале' in dff.columns and 'Цена_в_конце' in dff.columns:
        dff['Изменение_цены_%'] = (
            (dff['Цена_в_конце'] - dff['Цена_в_начале']) / dff['Цена_в_начале']
        ).round(4)

    if 'Средний_остаток' in dff.columns:
        dff['Оборачиваемость'] = (dff['Всего_продано'] / dff['Средний_остаток']).round(2)

    # Переименование колонок
    dff = dff.rename(columns={
        'Дней_продаж': 'Количество раз продаж',
        'Дней_в_наличии': 'Количество раз в наличии'
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        dff.to_excel(writer, index=False, sheet_name="Топ_пополнения")
        format_excel(dff, writer, sheet_name="Топ_пополнения")
    output.seek(0)

    return dcc.send_bytes(output.getvalue(), filename=f"топ_{top_n}_пополнения.xlsx")

HEIGHT_PER_BAR = 25  # Высота одной строки (можно подкорректировать)
MAX_CONTAINER_HEIGHT = 700  # Максимальная высота контейнера в px (как в layout)
@app.callback(
    Output('graph-top-fast', 'figure'),
    Input('sklad-filter', 'value'),
    Input('top-n-selector', 'value'),
)
def update_top_fast(selected_sklad, top_n):
    if not selected_sklad:
        return go.Figure()

    dff = fast_grouped[fast_grouped['Склад'].isin(selected_sklad)]
    dff = dff.sort_values('Всего_продано', ascending=False).head(top_n)

    graph_height = HEIGHT_PER_BAR * len(dff)
    graph_height = min(graph_height, MAX_CONTAINER_HEIGHT)

    fig = px.bar(
        dff,
        y='Номенклатура',
        x='Всего_продано',
        color='Склад',
        orientation='h',
        height=graph_height,
        title=f'Топ-{top_n} самых ходовых товаров'
    )

    fig.update_layout(
        yaxis={
            'categoryorder': 'array',
            'categoryarray': dff['Номенклатура'][::-1]  # переворачиваем порядок
        },
        template='plotly_white',
        margin=dict(l=250),
    )
    return fig


@app.callback(
    Output('graph-top-restock', 'figure'),
    Input('sklad-filter', 'value'),
    Input('top-n-selector-restock', 'value'),
)
def update_top_restock(selected_sklads, top_n):
    if not selected_sklads:
        return go.Figure()

    dff = restock_grouped[restock_grouped['Склад'].isin(selected_sklads)]
    dff = dff.sort_values('Всего_пополнено', ascending=False).head(top_n)

    graph_height = HEIGHT_PER_BAR * len(dff)
    graph_height = min(graph_height, MAX_CONTAINER_HEIGHT)

    fig = px.bar(
        dff,
        y='Номенклатура',
        x='Всего_пополнено',
        color='Склад',
        orientation='h',
        height=graph_height,
        title=f'Топ-{top_n} товаров по пополнениям'
    )

    fig.update_layout(
        yaxis={
            'categoryorder': 'array',
            'categoryarray': dff['Номенклатура'][::-1]  # переворачиваем порядок
        },
        template='plotly_white',
        margin=dict(l=250),
    )
    return fig

@app.callback(
    Output("peak-nom-filter", "options"),
    Input("peak-sklad-filter", "value"),
    Input("peak-article-filter", "value")
)
def update_nom_options(selected_sklad, selected_article):
    if not selected_sklad and not selected_article:
        return []

    dff = df_peaks.copy()
    if selected_sklad:
        dff = dff[dff["Склад"] == selected_sklad]
    if selected_article:
        dff = dff[dff["Артикул"] == selected_article]

    return [{"label": nom, "value": nom} for nom in sorted(dff["Номенклатура"].unique())]

@app.callback(
    Output('graph-peaks', 'figure'),
    Input('peak-sklad-filter', 'value'),
    Input('peak-article-filter', 'value'),
    Input('peak-nom-filter', 'value'),
)
def update_peaks_graph(sklad, article, nom):
    dff = df_peaks.copy()
    if sklad:
        dff = dff[dff['Склад'] == sklad]
    if article:
        dff = dff[dff['Артикул'] == article]
    if nom:
        dff = dff[dff['Номенклатура'] == nom]

    if dff.empty:
        return go.Figure()

    dff = dff.sort_values('Дата').tail(200)  # ограничение последних 200 строк

    fig = go.Figure()

    for sklad_name, group in dff.groupby('Склад'):
        fig.add_trace(go.Scatter(
            x=group['Дата'],
            y=group['Всего_продано'],
            mode='markers+lines',
            name=f'Продано - {sklad_name}',
            hovertemplate='Дата: %{x}<br>Продано: %{y}<br>Артикул: %{customdata[0]}<br>Номенклатура: %{customdata[1]}<extra></extra>',
            customdata=group[['Артикул', 'Номенклатура']],
            yaxis='y1',
        ))

        fig.add_trace(go.Scatter(
            x=group['Дата'],
            y=group['Средняя_цена'],
            mode='lines+markers',
            name=f'Средняя цена - {sklad_name}',
            line=dict(dash='dot'),
            hovertemplate='Дата: %{x}<br>Средняя цена: %{y}<extra></extra>',
            yaxis='y2',
        ))

        fig.add_trace(go.Scatter(
            x=group['Дата'],
            y=group['Изменение_цены_%'],
            mode='lines+markers',
            name=f'Изменение цены % - {sklad_name}',
            line=dict(dash='dash'),
            hovertemplate='Дата: %{x}<br>Изменение цены %: %{y}<extra></extra>',
            yaxis='y3',
        ))

    fig.update_layout(
        title='Всплески продаж и динамика цен',
        xaxis=dict(title='Дата'),
        yaxis=dict(title='Продано', side='left', showgrid=False, zeroline=False),
        yaxis2=dict(title='Средняя цена', overlaying='y', side='right', showgrid=False, zeroline=False, position=0.95),
        yaxis3=dict(title='Изменение цены %', overlaying='y', side='right', showgrid=False, zeroline=False,
                    position=1.0, anchor='free'),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=40, r=80, t=80, b=40),
        template='plotly_white'
    )
    return fig

@app.callback(
    Output("download-peaks-xlsx", "data"),
    Input("btn-download-peaks", "n_clicks"),
    State("peak-sklad-filter", "value"),
    State("peak-article-filter", "value"),
    State("peak-nom-filter", "value"),
    prevent_initial_call=True,
)
def download_peaks_excel(n_clicks, sklad, article, nom):
    dff = df_peaks.copy()
    if sklad:
        dff = dff[dff['Склад'] == sklad]
    if article:
        dff = dff[dff['Артикул'] == article]
    if nom:
        dff = dff[dff['Номенклатура'] == nom]

    if dff.empty:
        return dash.no_update

    # Добавим столбец с оборачиваемостью (если нет - считаем как пример)
    # Например: Оборачиваемость = Всего_продано / Среднее количество на складе (пример)
    # Здесь подставь свою логику, если нужно
    if 'Оборачиваемость' not in dff.columns:
        dff['Оборачиваемость'] = dff['Всего_продано'] / 10  # пример

    import io
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        dff.to_excel(writer, index=False, sheet_name='Всплески_продаж')
    output.seek(0)

    return dcc.send_bytes(output.read(), filename="всплески_продаж.xlsx")

@app.callback(
    Output("download-2025-xlsx", "data"),
    Input("download-2025-btn", "n_clicks"),
    State("sklad-2025-filter", "value"),
    State("article-2025-filter", "value"),
    State("month-2025-filter", "value"),
    prevent_initial_call=True,
)
def download_2025_excel(n_clicks, sklad, article, month):
    try:
        dff = df_2025.copy()
        dff.columns = [col.strip() for col in dff.columns]
        print(f"[download_2025_excel] Колонки df_2025: {dff.columns.tolist()}")

        col_map = {
            "Склад": "Склад",
            "Артикул": "Артикул_товар",
            "Дата": "Дата",
            "Остаток": "Остаток",
            "Цена": "Цена"
        }

        # --- обязательные колонки ---
        for r in col_map.values():
            if r not in dff.columns:
                print(f"[download_2025_excel] ERROR: нет колонки '{r}'")
                return dash.no_update

        # --- фильтры ---
        if sklad:
            if isinstance(sklad, list):
                dff = dff[dff[col_map["Склад"]].isin(sklad)]
            else:
                dff = dff[dff[col_map["Склад"]] == sklad]

        if article:
            dff = dff[dff[col_map["Артикул"]] == article]

        if month:
            dff = dff[dff[col_map["Дата"]].dt.month == int(month)]

        if dff.empty:
            print("[download_2025_excel] INFO: пустой датафрейм после фильтров")
            return dash.no_update

        # --- расчет Продано и Пополнено ---
        sort_cols = [col_map["Склад"], col_map["Артикул"], col_map["Дата"]]
        dff = dff.sort_values(sort_cols).reset_index(drop=True)

        dff["diff"] = dff.groupby([col_map["Склад"], col_map["Артикул"]])[col_map["Остаток"]].diff().fillna(0)
        # Продано: только отрицательные изменения
        dff["Продано"] = (-dff["diff"]).clip(lower=0)
        # Пополнено: только положительные изменения
        dff["Пополнено"] = dff["diff"].clip(lower=0)
        dff.drop(columns=["diff"], inplace=True)

        # --- подготовка Excel ---
        import io
        import xlsxwriter
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            for skl in dff[col_map["Склад"]].unique():
                dff_skl = dff[dff[col_map["Склад"]] == skl]
                cols_to_export = [
                    col_map["Дата"], col_map["Склад"], col_map["Артикул"],
                    col_map["Остаток"], "Продано", "Пополнено", col_map["Цена"]
                ]
                dff_skl_export = dff_skl[cols_to_export].copy()
                # цена с ₽
                dff_skl_export[col_map["Цена"]] = dff_skl_export[col_map["Цена"]].apply(lambda x: f"{x:.2f} ₽")

                dff_skl_export.to_excel(writer, index=False, sheet_name=str(skl)[:31])
                # автоширина колонок
                worksheet = writer.sheets[str(skl)[:31]]
                for i, col in enumerate(dff_skl_export.columns):
                    max_len = max(dff_skl_export[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.set_column(i, i, max_len)

        output.seek(0)
        return dcc.send_bytes(output.read(), filename="данные_2025.xlsx")

    except Exception as e:
        print(f"[download_2025_excel] ERROR: {e}", e.__class__)
        return dash.no_update

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))  # Используем порт из переменной окружения или 10000 по умолчанию
    app.run_server(debug=False, host='0.0.0.0', port=port)
