import os
import re
import zipfile
import logging
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from git import Repo  # pip install gitpython

# === Пути ===
DATA_DIR = "data"
ARCHIVE_PATH = os.path.join(DATA_DIR, "aggregated.zip")
UPLOADS_PATH = os.path.join(DATA_DIR, "new_uploads")

# === Вспомогательные функции ===
def safe_filename(s: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", str(s))

def detect_sklad_from_filename(file_path: str) -> str:
    name = os.path.basename(file_path).lower()
    if "моск" in name:
        return "Москва"
    elif "хаб" in name or "khab" in name:
        return "Хабаровск"
    else:
        return "Неизвестно"

def parse_date_from_cell(cell_value, file_path):
    if cell_value is None:
        date = datetime.fromtimestamp(os.path.getctime(file_path))
        logging.warning(f"[parse_date_from_cell] Дата в файле {file_path} отсутствует, подставлена дата создания {date}")
        return date
    date_str = str(cell_value).strip()
    date = pd.to_datetime(date_str, errors="coerce", dayfirst=True)
    if pd.isna(date):
        date = datetime.fromtimestamp(os.path.getctime(file_path))
        logging.warning(f"[parse_date_from_cell] Дата '{date_str}' в файле {file_path} не распознана, подставлена дата создания {date}")
    return date

# === Чтение Excel ===
def read_excel_file(file_path, sklad_name="auto"):
    logging.info(f"[read_excel_file] Чтение файла: {file_path}")
    extension = os.path.splitext(file_path)[1].lower()
    data = []

    try:
        if extension == ".xls":
            df_raw = pd.read_excel(file_path, header=None, engine="xlrd")
            date_cell = df_raw.iloc[1, 0]
            date = parse_date_from_cell(date_cell, file_path)
            for idx in range(4, len(df_raw)):
                row = df_raw.iloc[idx]
                if pd.isna(row[5]):
                    continue
                data.append({
                    "Дата": date,
                    "Номенклатура": row[1],
                    "Остаток": row[2] or 0,
                    "Цена": row[3] or 0,
                    "Производитель": row[4],
                    "Артикул": str(row[5]).strip(),
                })

        elif extension == ".xlsx":
            wb = load_workbook(filename=file_path, data_only=True)
            ws = wb.active
            date_cell = ws["A2"].value
            date = parse_date_from_cell(date_cell, file_path)
            for i in range(5, ws.max_row + 1):
                article = ws[f"F{i}"].value
                if article is None:
                    continue
                data.append({
                    "Дата": date,
                    "Номенклатура": ws[f"B{i}"].value,
                    "Остаток": ws[f"C{i}"].value or 0,
                    "Цена": ws[f"D{i}"].value or 0,
                    "Производитель": ws[f"E{i}"].value,
                    "Артикул": str(article).strip(),
                })

        df = pd.DataFrame(data)
        if df.empty:
            logging.warning(f"[read_excel_file] Файл {file_path} пуст после обработки")
            return None

        if sklad_name == "auto":
            df["Склад"] = detect_sklad_from_filename(file_path)
        else:
            df["Склад"] = sklad_name

        logging.info(f"[read_excel_file] Успешно загружено строк: {len(df)} из {file_path}")
        return df

    except Exception as e:
        logging.error(f"[read_excel_file] Ошибка при чтении {file_path}: {e}", exc_info=True)
        return None

# === Обработка нового файла ===
def process_new_file(file_path, output_folder=UPLOADS_PATH, repo_path="."):
    logging.info(f"[process_new_file] Начата обработка {file_path}")
    df_new = read_excel_file(file_path, sklad_name="auto")

    if df_new is None or df_new.empty:
        msg = f"[process_new_file] Файл {file_path} пуст или не содержит данных"
        logging.warning(msg)
        return 0, msg

    required = ["Артикул", "Номенклатура", "Остаток", "Цена", "Дата", "Склад"]
    missing = [c for c in required if c not in df_new.columns]
    if missing:
        msg = f"[process_new_file] В файле {file_path} отсутствуют колонки: {missing}"
        logging.error(msg)
        return 0, msg

    added_rows = 0
    for (sklad, article), group in df_new.groupby(["Склад", "Артикул"]):
        folder = os.path.join(output_folder, safe_filename(sklad))
        os.makedirs(folder, exist_ok=True)
        out_file = os.path.join(folder, f"{safe_filename(article)}.csv")

        if os.path.exists(out_file):
            df_old = pd.read_csv(out_file)
            if "Дата" not in df_old.columns:
                logging.warning(f"[process_new_file] В старом CSV {out_file} нет колонки 'Дата'")
                continue
            max_date = pd.to_datetime(df_old["Дата"]).max()
            group = group[pd.to_datetime(group["Дата"]) > max_date]

            if not group.empty:
                group.to_csv(out_file, mode="a", header=False, index=False, encoding="utf-8-sig")
                added_rows += len(group)
                logging.info(f"[process_new_file] Добавлено {len(group)} строк в {out_file}")
        else:
            group.to_csv(out_file, index=False, encoding="utf-8-sig")
            added_rows += len(group)
            logging.info(f"[process_new_file] Создан новый файл {out_file}, строк: {len(group)}")

    if added_rows > 0:
        try:
            git_autocommit(repo_path, file_path, added_rows)
        except Exception as e:
            msg = f"[process_new_file] Данные добавлены ({added_rows}), но ошибка автокоммита: {e}"
            logging.error(msg, exc_info=True)
            return added_rows, msg

    logging.info(f"[process_new_file] Завершено. Добавлено строк: {added_rows}")
    return added_rows, None

# === Автокоммит ===
def git_autocommit(repo_path, file_path, added_rows):
    try:
        repo = Repo(repo_path)
        repo.git.add(A=True)
        commit_msg = f"Автодобавление новых данных: {os.path.basename(file_path)}, строк {added_rows}"
        repo.index.commit(commit_msg)
        origin = repo.remote(name="origin")
        origin.push()
        logging.info(f"[git_autocommit] Автокоммит выполнен: {commit_msg}")
    except Exception as e:
        logging.error(f"[git_autocommit] Ошибка автокоммита: {e}", exc_info=True)
        raise

# === Чтение архива aggregated.zip ===
def load_from_archive(archive_path=ARCHIVE_PATH):
    all_dfs = []
    if not os.path.exists(archive_path):
        logging.warning(f"[load_from_archive] Архив {archive_path} не найден")
        return pd.DataFrame()

    with zipfile.ZipFile(archive_path, "r") as z:
        for fname in z.namelist():
            if not fname.endswith(".csv"):
                continue
            with z.open(fname) as f:
                try:
                    df = pd.read_csv(f)
                    all_dfs.append(df)
                except Exception as e:
                    logging.error(f"[load_from_archive] Ошибка чтения {fname}: {e}")
    return pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

# === Чтение new_uploads ===
def load_from_new_uploads(uploads_path=UPLOADS_PATH):
    all_dfs = []
    if not os.path.exists(uploads_path):
        os.makedirs(uploads_path, exist_ok=True)
        return pd.DataFrame()

    for root, _, files in os.walk(uploads_path):
        for fname in files:
            if not fname.endswith(".csv"):
                continue
            fpath = os.path.join(root, fname)
            try:
                df = pd.read_csv(fpath)
                all_dfs.append(df)
            except Exception as e:
                logging.error(f"[load_from_new_uploads] Ошибка чтения {fpath}: {e}")
    return pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

# === Объединение архив + новые ===
def load_all_data():
    df_archive = load_from_archive()
    df_new = load_from_new_uploads()
    if df_archive.empty and df_new.empty:
        logging.warning("[load_all_data] Нет данных ни в архиве, ни в new_uploads")
        return pd.DataFrame()
    return pd.concat([df_archive, df_new], ignore_index=True, sort=False)
