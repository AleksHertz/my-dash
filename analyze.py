import os
import re

import pandas as pd
import glob
from openpyxl import load_workbook
from datetime import datetime
import plotly.express as px
import logging
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import numpy as np
import time
import functools

# === НАСТРОЙКА ЛОГИРОВАНИЯ ===
os.makedirs("логи", exist_ok=True)
logging.basicConfig(
    filename='логи/анализ_склада.log',
    level=logging.INFO,
    format='%(asctime)s — %(levelname)s — %(message)s'
)

def timing_decorator(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        duration = end - start
        msg = f"Время выполнения функции '{func.__name__}': {duration:.3f} секунд"
        print(msg)
        logging.info(msg)
        return result
    return wrapper

@timing_decorator
def parse_date_from_cell(cell_value, file_path):
    """
    Безопасно парсит дату из значения ячейки.
    Если дата не распознана, берёт дату создания файла.
    Логирует результат.
    """
    if cell_value is None:
        date = datetime.fromtimestamp(os.path.getctime(file_path))
        logging.warning(f'Дата в файле {file_path} отсутствует, подставлена дата создания файла: {date}')
        return date

    date_str = str(cell_value).strip()
    date = pd.to_datetime(date_str, errors='coerce', dayfirst=True)
    if pd.isna(date):
        date = datetime.fromtimestamp(os.path.getctime(file_path))
        logging.warning(f'Дата "{date_str}" в файле {file_path} не распознана, подставлена дата создания файла: {date}')
    else:
        logging.info(f'Дата из файла {file_path}: "{date_str}" распознана как {date}')
    return date

@timing_decorator
def read_excel_file(file_path, sklad_name):
    extension = os.path.splitext(file_path)[1].lower()
    data = []

    try:
        if extension == '.xls':
            df_raw = pd.read_excel(file_path, header=None, engine='xlrd')
            date_cell = df_raw.iloc[1, 0]
            date = parse_date_from_cell(date_cell, file_path)

            for idx in range(4, len(df_raw)):
                row = df_raw.iloc[idx]
                if pd.isna(row[5]):
                    continue
                data.append({
                    'Дата': date,
                    'Номенклатура': row[1],
                    'Количество': row[2] or 0,
                    'Цена': row[3] or 0,
                    'Производитель': row[4],
                    'Артикул': str(row[5]).strip(),
                    'Склад': sklad_name
                })

        elif extension == '.xlsx':
            wb = load_workbook(filename=file_path, data_only=True)
            ws = wb.active
            date_cell = ws['A2'].value
            date = parse_date_from_cell(date_cell, file_path)

            for i in range(5, ws.max_row + 1):
                article = ws[f'F{i}'].value
                if article is None:
                    continue
                data.append({
                    'Дата': date,
                    'Номенклатура': ws[f'B{i}'].value,
                    'Количество': ws[f'C{i}'].value or 0,
                    'Цена': ws[f'D{i}'].value or 0,
                    'Производитель': ws[f'E{i}'].value,
                    'Артикул': str(article).strip(),
                    'Склад': sklad_name
                })

        logging.info(f'Успешно прочитан файл: {file_path}')
        return pd.DataFrame(data)

    except Exception as e:
        logging.error(f'Ошибка при чтении файла {file_path}: {e}')
        return None
@timing_decorator
def process_folder(folder_path, sklad_name):
    files = sorted(glob.glob(os.path.join(folder_path, '*.xls*')))
    dfs = []
    for file in files:
        df = read_excel_file(file, sklad_name)
        if df is not None and not df.empty:
            dfs.append(df)
    if not dfs:
        logging.warning(f'Нет данных в папке: {folder_path}')
        return None
    return pd.concat(dfs, ignore_index=True)
@timing_decorator
def generate_daily_sales_file(df_all: pd.DataFrame, output_path: str = 'итог_дневные_продажи.csv'):
    try:
        if 'Дата' not in df_all.columns:
            logging.error("Колонка 'Дата' отсутствует в данных — невозможно сформировать файл ежедневных продаж.")
            return

        if 'Количество' not in df_all.columns:
            logging.error("Колонка 'Количество' отсутствует в данных — невозможно сформировать файл ежедневных продаж.")
            return

        if 'Цена' not in df_all.columns:
            logging.error("Колонка 'Цена' отсутствует в данных — невозможно сформировать файл с ценами.")
            return

        # Агрегация количества и цены (цена — первая за день)
        df_daily = (
            df_all
            .sort_values('Дата')  # чтобы first() работал корректно
            .groupby(['Дата', 'Артикул', 'Склад'], as_index=False)
            .agg({
                'Количество': 'sum',
                'Цена': 'first'
            })
            .rename(columns={'Количество': 'Всего_продано', 'Цена': 'Цена_в_начале_дня'})
        )

        # Сортировка
        df_daily.sort_values(['Дата', 'Склад', 'Артикул'], inplace=True)

        # ✅ Сохранение в CSV
        df_daily.to_csv(output_path, index=False, encoding='utf-8-sig')
        logging.info(f"📁 CSV-файл с дневными продажами и ценами сохранён: {output_path}")

    except Exception as e:
        logging.error(f"❌ Ошибка при создании файла с дневными продажами: {e}")


@timing_decorator
def analyze_with_restock_vectorized_monthly(df_all):
    # --- Подготовка и агрегация данных ---
    df_all['Дата'] = pd.to_datetime(df_all['Дата'], format='%d-%m-%Y', errors='coerce')

    df_daily = df_all.groupby(['Артикул', 'Склад', 'Дата'], as_index=False).agg({
        'Количество': 'first',
        'Цена': 'first',
        'Номенклатура': 'first',
        'Производитель': 'first'
    })

    df_daily = df_daily.sort_values(['Артикул', 'Склад', 'Дата']).copy()

    df_daily['Год'] = df_daily['Дата'].dt.year
    df_daily['Месяц'] = df_daily['Дата'].dt.month

    # --- Вычисление продаж и пополнений ---
    df_daily['diff_qty'] = df_daily.groupby(['Артикул', 'Склад'])['Количество'].diff().fillna(0)
    df_daily['Продано'] = (-df_daily['diff_qty']).clip(lower=0)
    df_daily['Пополнение'] = df_daily['diff_qty'].clip(lower=0)

    # --- Поиск перемещений между складами ---
    restock_rows = df_daily[df_daily['diff_qty'] > 0][['Дата', 'Артикул', 'Склад', 'diff_qty']].copy()
    restock_rows.rename(columns={'Склад': 'Склад_куда', 'diff_qty': 'Кол-во'}, inplace=True)

    sold_rows = df_daily[df_daily['diff_qty'] < 0][['Дата', 'Артикул', 'Склад', 'diff_qty']].copy()
    sold_rows.rename(columns={'Склад': 'Склад_откуда', 'diff_qty': 'Кол-во'}, inplace=True)
    sold_rows['Кол-во'] = -sold_rows['Кол-во']

    merged_moves = pd.merge(
        restock_rows, sold_rows,
        on=['Дата', 'Артикул', 'Кол-во'], how='inner'
    )
    merged_moves = merged_moves[merged_moves['Склад_куда'] != merged_moves['Склад_откуда']]

    for _, row in merged_moves.iterrows():
        mask_restock = (
            (df_daily['Дата'] == row['Дата']) &
            (df_daily['Артикул'] == row['Артикул']) &
            (df_daily['Склад'] == row['Склад_куда']) &
            (df_daily['diff_qty'] == row['Кол-во'])
        )
        df_daily.loc[mask_restock, 'Пополнение'] = 0

        mask_sold = (
            (df_daily['Дата'] == row['Дата']) &
            (df_daily['Артикул'] == row['Артикул']) &
            (df_daily['Склад'] == row['Склад_откуда']) &
            (df_daily['diff_qty'] == -row['Кол-во'])
        )
        df_daily.loc[mask_sold, 'Продано'] = 0

    перемещения = merged_moves.rename(columns={
        'Дата': 'Дата',
        'Артикул': 'Артикул',
        'Склад_откуда': 'Склад_откуда',
        'Склад_куда': 'Склад_куда',
        'Кол-во': 'Кол-во'
    })[['Дата', 'Артикул', 'Склад_откуда', 'Склад_куда', 'Кол-во']].to_dict(orient='records')

    # --- Фильтр цен (оставил как в оригинале, закомментировано) ---
    df_unique_price = df_daily.groupby(['Артикул', 'Склад', 'Дата'], as_index=False)['Цена'].first()

    price_counts = df_unique_price.groupby([
        'Артикул', 'Склад',
        df_unique_price['Дата'].dt.year.rename('Год'),
        df_unique_price['Дата'].dt.month.rename('Месяц'),
        'Цена'
    ])['Дата'].nunique().reset_index(name='Дней_с_ценой')

    # price_counts = price_counts[price_counts['Дней_с_ценой'] >= 2]

    df_filtered_price = pd.merge(
        df_unique_price,
        price_counts[['Артикул', 'Склад', 'Год', 'Месяц', 'Цена']],
        on=['Артикул', 'Склад', 'Цена'],
        how='inner'
    )

    price_stats = df_unique_price.groupby(['Артикул', 'Склад', df_unique_price['Дата'].dt.year.rename('Год'),
                                           df_unique_price['Дата'].dt.month.rename('Месяц')]).agg(
        Мин_цена=('Цена', 'min'),
        Макс_цена=('Цена', 'max')
    ).reset_index()

    # --- Продажи по месяцам ---
    df_sales = df_daily.groupby(['Артикул', 'Склад', 'Год', 'Месяц']).agg(
        Номенклатура=('Номенклатура', 'first'),
        Производитель=('Производитель', 'first'),
        Всего_продано=('Продано', 'sum'),
        Всего_пополнено=('Пополнение', 'sum'),
        Дней_продаж=('Продано', lambda x: (x > 0).sum()),
        Средняя_цена=('Цена', 'mean'),
        Дней_в_наличии=('Количество', lambda x: (x > 0).sum()),
        Последний_остаток=('Количество', 'last'),
        Уникальных_дней=('Дата', 'nunique')
    ).reset_index()

    df_sales['Дней_в_наличии'] = df_sales[['Дней_в_наличии', 'Уникальных_дней']].min(axis=1)
    df_sales['Оборачиваемость'] = df_sales['Всего_продано'] / df_sales['Дней_продаж'].replace(0, 1)

    df_price = df_daily.sort_values('Дата').groupby(['Артикул', 'Склад', 'Год', 'Месяц']).agg(
        Цена_в_начале=('Цена', 'first'),
        Цена_в_конце=('Цена', 'last')
    ).reset_index()

    df_price['Изменение_цены_абс'] = df_price['Цена_в_конце'] - df_price['Цена_в_начале']
    df_price['Изменение_цены_%'] = ((df_price['Цена_в_конце'] / df_price['Цена_в_начале']) - 1) * 100
    df_price['Изменение_цены_%'] = df_price['Изменение_цены_%'].fillna(0)
    df_price['Изменение_цены_абс'] = df_price['Изменение_цены_абс'].fillna(0)

    df_sales = df_sales.merge(price_stats, on=['Артикул', 'Склад', 'Год', 'Месяц'], how='left')
    df_sales = df_sales.merge(df_price, on=['Артикул', 'Склад', 'Год', 'Месяц'], how='left')

    # --- Проверка артикула/номенклатуры на возможные подмены ---
    def normalize_article(article):
        if pd.isna(article):
            return article
        return article.replace('-', '').replace(' ', '').upper()

    def normalize_nomenclature(nom):
        if pd.isna(nom):
            return nom
        nom = nom.lower().replace('дефект', '').strip()
        return nom

    df_all['Артикул_норм'] = df_all['Артикул'].apply(normalize_article)
    df_all['Номенклатура_норм'] = df_all['Номенклатура'].apply(normalize_nomenclature)

    grouped = df_all.groupby('Номенклатура_норм')['Артикул_норм'].nunique().reset_index()
    problematic = grouped[grouped['Артикул_норм'] > 1]

    problematic_articles = df_all[df_all['Номенклатура_норм'].isin(problematic['Номенклатура_норм'])][
        ['Номенклатура', 'Артикул', 'Артикул_норм', 'Номенклатура_норм']].drop_duplicates().reset_index(drop=True)

    return df_sales, перемещения, problematic_articles

@timing_decorator
def normalize_article(article):
    # Простая нормализация артикула: удаляем дефисы, пробелы, подчёркивания и приводим к верхнему регистру
    if not isinstance(article, str):
        return article
    return article.replace('-', '').replace(' ', '').replace('_', '').upper()

def run_month_analysis():
    logging.info("🔍 Начало анализа месяца")

    df_moscow = process_folder('data/moscow', 'Москва')
    df_khabarovsk = process_folder('data/khabarovsk', 'Хабаровск')

    if df_moscow is None and df_khabarovsk is None:
        logging.error("❌ Нет данных для анализа")
        return

    df_all = pd.concat([df for df in [df_moscow, df_khabarovsk] if df is not None], ignore_index=True)
    df_all.dropna(subset=['Артикул'], inplace=True)

    # Приводим артикулы к строкам, убираем пробелы и нормализуем
    df_all['Артикул'] = df_all['Артикул'].astype(str).str.strip().apply(normalize_article)

    df_all['Дата'] = pd.to_datetime(df_all['Дата'], errors='coerce')

    df_result, перемещения, df_flags = analyze_with_restock_vectorized_monthly(df_all)

    # Переименование колонок для унификации
    rename_map = {
        'Количество': 'Всего_продано',
        'Продано': 'Всего_продано',
        'Всего продано': 'Всего_продано',
        'Пополнено': 'Всего_пополнено',
        'Дни продаж': 'Дней_продаж',
        'Дней продаж': 'Дней_продаж',
        'Цена в начале': 'Цена_в_начале',
        'Цена в конце': 'Цена_в_конце',
        'Дни в наличии': 'Дней_в_наличии',
    }
    df_result.rename(columns=rename_map, inplace=True)

    needed_cols = ['Всего_продано', 'Всего_пополнено', 'Дней_продаж',
                   'Средняя_цена', 'Мин_цена', 'Макс_цена', 'Цена_в_начале',
                   'Цена_в_конце', 'Дней_в_наличии']
    for col in needed_cols:
        if col not in df_result.columns:
            df_result[col] = 0 if 'цена' not in col.lower() else None

    if isinstance(df_flags, list):
        df_flags = pd.DataFrame(df_flags) if df_flags else pd.DataFrame()

    if not df_flags.empty and 'Дата' in df_flags.columns:
        df_flags['Дата'] = pd.to_datetime(df_flags['Дата'], errors='coerce').dt.strftime('%d/%m/%Y')

    df_result.to_excel('итог_по_месяцу.xlsx', index=False)
    df_flags.to_excel('фиксация_перемещений.xlsx', index=False)

    df_total = df_result.groupby(['Артикул', 'Склад']).agg(
        Номенклатура=('Номенклатура', 'first'),
        Производитель=('Производитель', 'first'),
        Всего_продано=('Всего_продано', 'sum'),
        Всего_пополнено=('Всего_пополнено', 'sum'),
        Дней_продаж=('Дней_продаж', 'sum'),
        Средняя_цена=('Средняя_цена', 'mean'),
        Мин_цена=('Мин_цена', 'min'),
        Макс_цена=('Макс_цена', 'max'),
        Цена_в_начале=('Цена_в_начале', 'first'),
        Цена_в_конце=('Цена_в_конце', 'last'),
        Дней_в_наличии=('Дней_в_наличии', 'sum')
    ).reset_index()

    df_total['Дней_продаж'] = df_total['Дней_продаж'].replace(0, 1)
    df_total['Оборачиваемость'] = df_total['Всего_продано'] / df_total['Дней_продаж']

    top_fast = df_total[df_total['Всего_продано'] > 0].sort_values('Оборачиваемость', ascending=False).head(1000)
    top_slow = df_total[df_total['Всего_продано'] == 0].sort_values('Дней_в_наличии', ascending=False).head(1000)
    top_restocked = df_total.sort_values('Всего_пополнено', ascending=False).head(1000)

    top_fast.to_excel('самые_ходовые.xlsx', index=False)
    top_slow.to_excel('залежалые.xlsx', index=False)
    top_restocked.to_excel('чаще_всего_пополнялись.xlsx', index=False)

    generate_daily_sales_file(df_all, output_path='итог_дневные_продажи.csv')

    logging.info("✅ Анализ месяца завершен")


if __name__ == '__main__':
    run_month_analysis()